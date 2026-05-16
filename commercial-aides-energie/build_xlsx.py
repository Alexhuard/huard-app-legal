"""
Génération du classeur Excel "Aides & Primes Energie - Groupe HUARD"
Outil commercial interne : recensement des dispositifs nationaux et regionaux
mobilisables pour les clients de Groupe HUARD (Electricite, CVC, Telecoms,
Informatique, Maintenance) - mai 2026.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.dimensions import ColumnDimension


OUT = "/home/user/huard-app-legal/commercial-aides-energie/Aides-Primes-Energie-HUARD-2026.xlsx"

# Charte couleur (proche de la charte HUARD vue dans la PP : #1F2F4D / #5D81A6)
NAVY = "1F2F4D"
BLUE = "5D81A6"
LIGHT = "E8EEF5"
ACCENT = "F4A261"
GREEN = "2A9D8F"
RED = "E76F51"
GREY = "5A6478"

thin = Side(style="thin", color="BFC9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def title_cell(ws, cell, text, size=18, color="FFFFFF", fill=NAVY):
    ws[cell] = text
    ws[cell].font = Font(name="Calibri", size=size, bold=True, color=color)
    ws[cell].fill = PatternFill("solid", fgColor=fill)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", indent=1)


def header_row(ws, row, headers, fill=BLUE, color="FFFFFF"):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color=color, size=11)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_data_row(ws, row, n_cols, zebra=False):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        c.font = Font(name="Calibri", size=10)
        if zebra:
            c.fill = PatternFill("solid", fgColor=LIGHT)


wb = Workbook()

# =========================================================================
# 1. PAGE D'ACCUEIL / SOMMAIRE
# =========================================================================
ws = wb.active
ws.title = "Sommaire"
ws.sheet_view.showGridLines = False

title_cell(ws, "B2", "GROUPE HUARD - AIDES & PRIMES ENERGIE 2026", size=20)
ws.merge_cells("B2:H2")
ws.row_dimensions[2].height = 38

ws["B3"] = "Outil commercial interne - Recensement des dispositifs mobilisables pour nos clients"
ws["B3"].font = Font(size=11, italic=True, color=GREY)
ws.merge_cells("B3:H3")

ws["B5"] = "Date de mise a jour"
ws["B5"].font = Font(bold=True, color=NAVY)
ws["C5"] = "16 mai 2026"

ws["B6"] = "Perimetre"
ws["B6"].font = Font(bold=True, color=NAVY)
ws["C6"] = "France entiere + focus Ile-de-France (siege Bievres 91)"

ws["B7"] = "Activites HUARD couvertes"
ws["B7"].font = Font(bold=True, color=NAVY)
ws["C7"] = "Electricite (courants forts/faibles) - CVC - Telecoms - Informatique - Maintenance multi-technique"

ws["B8"] = "Cibles clients"
ws["B8"].font = Font(bold=True, color=NAVY)
ws["C8"] = "Tertiaire prive, industrie, copropriete, collectivites, bailleurs sociaux, logement individuel et collectif"

ws["B9"] = "Auteur"
ws["B9"].font = Font(bold=True, color=NAVY)
ws["C9"] = "Direction commerciale - Groupe HUARD"

# Sommaire
title_cell(ws, "B11", "SOMMAIRE", size=14, fill=BLUE)
ws.merge_cells("B11:H11")

sommaire = [
    ("01", "Tableau complet - toutes les aides", "Base de donnees principale (37 dispositifs) avec criteres, montants, levier commercial"),
    ("02", "Vue par activite HUARD", "Filtrage des dispositifs par metier : ELEC / CVC / TELECOM-INFO / MAINTENANCE"),
    ("03", "Eligibilite clients", "Grille de qualification rapide pour le commercial : qui peut beneficier de quoi"),
    ("04", "Argumentaire commercial", "Pitch type, objections, cumul des aides, exemples chiffres"),
    ("05", "CEE - Fiches d'operation cles", "Detail des fiches CEE BAT, IND et BAR les plus rentables pour nos chantiers"),
    ("06", "Decret tertiaire (DEET)", "Obligations reglementaires - opportunite commerciale n 1 du tertiaire"),
    ("07", "Aides Ile-de-France", "Specifique Region IDF, MGP, ADEME IDF, Energies POSIT'IF"),
    ("08", "Sources & liens officiels", "URL des dispositifs, contacts utiles, references reglementaires"),
]
header_row(ws, 12, ["#", "Onglet", "Contenu"], fill=NAVY)
ws.column_dimensions["B"].width = 6
ws.column_dimensions["C"].width = 38
ws.column_dimensions["D"].width = 90

for i, (n, t, d) in enumerate(sommaire, start=13):
    ws.cell(row=i, column=2, value=n).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=i, column=3, value=t).font = Font(bold=True, color=NAVY)
    ws.cell(row=i, column=4, value=d).alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(2, 5):
        ws.cell(row=i, column=col).border = border
    ws.row_dimensions[i].height = 28

# Avertissement
ws["B23"] = "AVERTISSEMENT"
ws["B23"].font = Font(bold=True, color="FFFFFF")
ws["B23"].fill = PatternFill("solid", fgColor=RED)
ws.merge_cells("B23:H23")

avert = (
    "Les montants, plafonds et conditions d'eligibilite sont ceux en vigueur au 16/05/2026. "
    "Les baremes CEE (bonifications, coups de pouce), les arretes tarifaires PV (S21/S26) et les enveloppes ADEME "
    "evoluent chaque trimestre voire chaque mois. AVANT TOUT ENGAGEMENT COMMERCIAL CHIFFRE, valider les montants "
    "actualises aupres de : (1) le mandataire CEE partenaire, (2) ADEME / Bpifrance pour les subventions, "
    "(3) la Region IDF pour les aides regionales. Document non contractuel - usage interne strict."
)
ws["B24"] = avert
ws["B24"].font = Font(size=10, italic=True, color=GREY)
ws["B24"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B24:H27")
ws.row_dimensions[24].height = 25
for r in range(24, 28):
    ws.row_dimensions[r].height = 22

# =========================================================================
# 2. TABLEAU COMPLET DES AIDES
# =========================================================================
ws2 = wb.create_sheet("01 - Tableau complet")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "B4"

title_cell(ws2, "A1", "RECENSEMENT COMPLET DES AIDES ENERGIE - MAI 2026", size=16)
ws2.merge_cells("A1:T1")
ws2.row_dimensions[1].height = 32

ws2["A2"] = "37 dispositifs - tri par categorie - filtres actives sur la ligne 3"
ws2["A2"].font = Font(italic=True, color=GREY, size=10)
ws2.merge_cells("A2:T2")

headers = [
    "N",
    "Categorie",
    "Nom du dispositif",
    "Type",
    "Financeur",
    "Beneficiaires eligibles",
    "Activite HUARD",
    "Travaux / equipements eligibles",
    "Montant min",
    "Montant max / Plafond",
    "Mode de calcul",
    "Criteres d'eligibilite cles",
    "Certifications exigees",
    "Cumul possible avec",
    "Qui depose le dossier",
    "Calendrier / validite",
    "Reference texte",
    "Lien officiel",
    "Levier commercial HUARD",
    "Priorite",
]
header_row(ws2, 3, headers)
ws2.row_dimensions[3].height = 50

# Donnees - chaque tuple = 1 dispositif
# (categorie, nom, type, financeur, beneficiaires, activite, travaux, min, max, calcul, criteres, certifs, cumul, qui_depose, calendrier, ref, lien, levier, priorite)

D = []

# --- CEE - certificats d'economie d'energie ---
D += [
    ("CEE - Tertiaire", "CEE BAT-EQ-127 - Luminaire LED tertiaire", "CEE fiche standardisee", "Obliges CEE (Total, EDF, Engie, Auchan, mandataires)",
     "Tous proprietaires/exploitants de batiments tertiaires (bureaux, commerces, enseignement, sante, hotellerie, logistique)",
     "ELECTRICITE",
     "Remplacement de luminaires par modules LED neufs avec systeme de gestion (detection presence/luminosite). Marquage CE, efficacite >= 130 lm/W.",
     "Variable selon kWh cumac generes (typique 4 a 12 EUR / luminaire)",
     "Pas de plafond legal - selon volume et marche CEE (cours kWh cumac ~7-9 EUR/MWh cumac en 2026)",
     "Forfait kWh cumac x zone climatique x duree d'usage / cours du marche CEE",
     "Travaux dans batiment existant > 2 ans - signature devis APRES engagement CEE - facture detaillee - PV de pose",
     "Aucune RGE exigee en tertiaire pour cette fiche (RGE non obligatoire BAT)",
     "Coup de pouce eclairage (si actif) - aides regionales - suramortissement",
     "HUARD (en partenariat avec mandataire CEE) ou client direct",
     "Dispositif perenne - 5eme periode CEE jusqu'au 31/12/2026, 6eme periode en preparation",
     "Arrete du 22/12/2014 modifie - fiche BAT-EQ-127 v.A35-5",
     "https://www.ecologie.gouv.fr/operations-standardisees-deconomies-denergie",
     "ARGUMENT N 1 ELEC : 'votre relamping LED finance a 30-70% par les CEE'. Cycle court, ROI < 3 ans, recurrent (relamping cyclique).",
     "Haute"),

    ("CEE - Tertiaire", "CEE BAT-EQ-133 - Systeme de gestion de l'eclairage", "CEE fiche standardisee", "Obliges CEE",
     "Tertiaire - parkings couverts, circulations, sanitaires, bureaux",
     "ELECTRICITE",
     "Detecteurs de presence/luminosite, gradation automatique, gestion centralisee KNX/DALI",
     "Variable (~3-6 EUR/m2 traite)", "Pas de plafond - selon kWh cumac",
     "Forfait par m2 zone gestion x DJU x cours CEE",
     "Pose dans batiment existant - asservissement effectif a la presence et a l'eclairement",
     "Aucune RGE obligatoire", "BAT-EQ-127 (relamping LED), GTB",
     "HUARD + mandataire CEE", "Periode CEE en cours",
     "Fiche CEE BAT-EQ-133", "https://www.ecologie.gouv.fr/operations-standardisees-deconomies-denergie",
     "Couple avec BAT-EQ-127 = offre globale eclairage clef en main. Marge sur produits + prime CEE.",
     "Haute"),

    ("CEE - Tertiaire", "CEE BAT-TH-116 - Systeme de GTB (Gestion Technique du Batiment)", "CEE fiche standardisee", "Obliges CEE",
     "Tertiaire > 1000 m2 (cible decret BACS) - tous segments",
     "ELECTRICITE / CVC / MAINTENANCE",
     "GTB classe A ou B (norme NF EN ISO 52120-1) pilotant CVC, eclairage, ECS, stores. Compatible decret BACS.",
     "Tres eleve - typique 8 a 35 EUR/m2 traite",
     "Souvent > 50 000 EUR pour un batiment de 5000 m2",
     "Forfait par m2 chauffe x classe GTB (A ou B) x usage x cours CEE",
     "Batiment > 2 ans - GTB classe A ou B verifiee - asservissement effectif - releve consommation",
     "Aucune RGE obligatoire", "Decret tertiaire (DEET), Coup de pouce pilotage, ADEME ACTEE",
     "HUARD (integrateur) + mandataire CEE",
     "Fiche bonifiee jusque fin 2026 - revision attendue 7eme periode CEE",
     "Fiche CEE BAT-TH-116 + decret BACS (decret n 2020-887)",
     "https://www.ecologie.gouv.fr/decret-bacs",
     "OPPORTUNITE MAJEURE : DECRET BACS rend obligatoire la GTB pour tertiaire avec systeme CVC > 290 kW (et 70 kW en 2027). Marche captif.",
     "Tres haute"),

    ("CEE - Tertiaire", "CEE BAT-TH-104 - Pompe a chaleur (PAC) air/eau, eau/eau ou eau glycolee/eau", "CEE fiche standardisee", "Obliges CEE",
     "Tertiaire et collectif - tous segments",
     "CVC",
     "Installation d'une PAC en remplacement d'un systeme existant. SCOP minimal selon temperature de depart.",
     "Variable - typique 1500 a 8000 EUR par PAC",
     "Variable selon puissance et kWh cumac",
     "Forfait kW x SCOP x zone climatique x cours CEE",
     "PAC neuve - SCOP >= 3,4 (basse T) ou 3,2 (moyenne T) - depose chaudiere fioul/gaz (bonifie si fioul)",
     "Qualification RGE QualiPAC ou equivalent recommandee (selon donneur d'ordre)",
     "Coup de pouce Chauffage tertiaire, Fonds Chaleur (geothermie), aides regionales",
     "HUARD (RGE) + mandataire CEE",
     "En vigueur - bonification renforcee sur sortie fioul jusque 31/12/2026",
     "Fiche CEE BAT-TH-104",
     "https://www.ecologie.gouv.fr/operations-standardisees-deconomies-denergie",
     "Cible : copropriete chauffage collectif fioul/gaz, batiments tertiaires anciens. Offre 'CVC decarbonee clef en main'.",
     "Tres haute"),

    ("CEE - Tertiaire", "CEE BAT-TH-127 - Raccordement reseau de chaleur EnR&R", "CEE fiche standardisee", "Obliges CEE",
     "Tertiaire et copro - dans communes desservies par un reseau de chaleur > 50% EnR&R",
     "CVC",
     "Sous-station de raccordement + travaux primaires",
     "1000 a 5000 EUR / logement equivalent", "Variable selon volume",
     "Forfait par logement equivalent ou par m2 chauffe x cours CEE",
     "Batiment existant > 2 ans - raccordement effectif - taux EnR&R reseau > 50%",
     "Aucune RGE obligatoire",
     "Coup de pouce raccordement, Fonds Chaleur, MaPrimeRenov Copropriete",
     "HUARD + mandataire CEE", "Bonification active sur reseaux EnR > 50%",
     "Fiche CEE BAT-TH-127", "https://www.ecologie.gouv.fr/operations-standardisees-deconomies-denergie",
     "IDF : plus de 100 reseaux de chaleur - argument fort copro / tertiaire en centres urbains.",
     "Haute"),

    ("CEE - Tertiaire", "CEE BAT-TH-146 - Calorifugeage de points singuliers", "CEE fiche standardisee", "Obliges CEE",
     "Tertiaire, collectif, industrie",
     "CVC / MAINTENANCE",
     "Calorifugeage des points singuliers de reseau (vannes, brides, pompes) en chaufferie",
     "100 a 600 EUR / point singulier", "Selon nombre de points",
     "Forfait par point x cours CEE",
     "Reseau > 50 degC - epaisseur isolant conforme - pose dans existant",
     "Aucune RGE obligatoire", "BAT-TH-145, contrats P2/P3",
     "HUARD (maintenance) + mandataire CEE", "En vigueur",
     "Fiche CEE BAT-TH-146",
     "https://www.ecologie.gouv.fr/operations-standardisees-deconomies-denergie",
     "Action recurrente facile a vendre en contrat maintenance - financement quasi 100% par CEE.",
     "Moyenne"),

    ("CEE - Tertiaire", "CEE BAT-TH-141 - Systeme de regulation par programmation horaire", "CEE fiche standardisee", "Obliges CEE",
     "Tertiaire",
     "CVC / ELECTRICITE",
     "Horloge / programmateur sur installation de chauffage ou ECS",
     "150 a 800 EUR / installation", "Selon kWh cumac",
     "Forfait x cours CEE", "Pose dans batiment > 2 ans - regulation effective",
     "Aucune RGE", "GTB (BAT-TH-116)", "HUARD + mandataire", "En vigueur",
     "Fiche CEE BAT-TH-141", "https://www.ecologie.gouv.fr/operations-standardisees-deconomies-denergie",
     "Action low-cost a integrer systematiquement dans toute prestation CVC.",
     "Moyenne"),

    ("CEE - Tertiaire", "CEE BAT-TH-155 / 125 - Ventilation double flux haute efficacite", "CEE fiche standardisee", "Obliges CEE",
     "Tertiaire",
     "CVC",
     "VMC double flux avec echangeur a haute efficacite (>= 75%)",
     "Variable", "Selon kWh cumac",
     "Forfait par m3/h x cours CEE",
     "Echangeur >= 75% - moteur basse conso - filtre F7 minimum",
     "Aucune RGE", "BAT-TH-116 (GTB), audit energetique",
     "HUARD + mandataire", "En vigueur",
     "Fiche CEE BAT-TH-155", "https://www.ecologie.gouv.fr/operations-standardisees-deconomies-denergie",
     "QAI + economies = double argument post-Covid, fort en bureaux/ecoles.",
     "Moyenne"),

    ("CEE - Industrie", "CEE IND-UT-103 - Variateur electronique de vitesse (VEV)", "CEE fiche standardisee", "Obliges CEE",
     "Sites industriels et tertiaires avec moteurs asynchrones",
     "ELECTRICITE / MAINTENANCE",
     "Pose VEV sur moteur asynchrone alimentant pompe/ventilateur/compresseur",
     "300 a 3000 EUR / moteur", "Selon kW",
     "Forfait par kW x cours CEE",
     "Moteur > 0,75 kW - charge variable - pose neuve",
     "Aucune RGE", "Toutes aides decarbonation industrie",
     "HUARD + mandataire", "En vigueur",
     "Fiche CEE IND-UT-103", "https://www.ecologie.gouv.fr/operations-standardisees-deconomies-denergie",
     "Forte rentabilite client - argument 'efficacite + CEE finance >50%'.",
     "Haute"),

    ("CEE - Industrie", "CEE IND-UT-117 - Recuperation de chaleur sur compresseur d'air", "CEE fiche standardisee", "Obliges CEE",
     "Sites industriels",
     "CVC / MAINTENANCE",
     "Recuperation chaleur fatale des compresseurs d'air vers ECS ou chauffage",
     "Variable - typique 3000 a 15000 EUR", "Selon kWh cumac",
     "Forfait par kW thermique recupere", "Compresseur > 10 kW - usage continu",
     "Aucune RGE", "Fonds Chaleur (chaleur fatale)", "HUARD + mandataire",
     "En vigueur", "Fiche CEE IND-UT-117",
     "https://www.ecologie.gouv.fr/operations-standardisees-deconomies-denergie",
     "Cible PME industrielles - couple Diag Eco-Flux + CEE + Fonds Chaleur.",
     "Moyenne"),
]

# --- Coups de pouce CEE bonifies ---
D += [
    ("Coup de pouce CEE", "Coup de pouce 'Chauffage tertiaire'", "CEE bonifie", "Obliges CEE signataires de la charte",
     "Batiments tertiaires - tous segments",
     "CVC",
     "Remplacement chaudiere charbon/fioul/gaz par : raccordement reseau de chaleur EnR, PAC, chaudiere biomasse",
     "5x a 10x les CEE standards", "Selon volume",
     "Bonification des kWh cumac x facteur 2 a 4 selon energie sortie",
     "Sortie fioul/charbon = bonification max - engagement avant fin charte",
     "Aucune RGE specifique (tertiaire)",
     "Fonds Chaleur, aides regionales, MaPrimeRenov Copro",
     "HUARD + mandataire signataire charte",
     "Operation engagee avant le 31/12/2026 (charte courante)",
     "Arrete coup de pouce - charte sectorielle",
     "https://www.ecologie.gouv.fr/coup-pouce-economies-denergie",
     "Argument 'derniere ligne droite avant fin du coup de pouce' - urgence commerciale forte.",
     "Tres haute"),

    ("Coup de pouce CEE", "Coup de pouce 'Pilotage connecte du chauffage'", "CEE bonifie", "Obliges CEE",
     "Tertiaire et residentiel - systemes de chauffage central",
     "CVC / ELECTRICITE",
     "Equipement de regulation centralise type GTB classe A/B ou thermostat connecte",
     "Bonification ~ x2 vs CEE standard", "Selon batiment",
     "Bonification kWh cumac", "Pose et asservissement effectif",
     "Aucune RGE", "GTB BAT-TH-116, BAT-TH-141",
     "HUARD + mandataire", "En vigueur",
     "Arrete coup de pouce pilotage", "https://www.ecologie.gouv.fr/coup-pouce-economies-denergie",
     "Couple parfait avec offre 'maintenance connectee' HUARD.",
     "Haute"),

    ("Coup de pouce CEE", "Coup de pouce 'Renovation performante de batiment tertiaire'", "CEE bonifie", "Obliges CEE",
     "Tertiaire - operations globales avec gain >= 30% energie finale",
     "ELECTRICITE / CVC / MAINTENANCE",
     "Bouquet de travaux : isolation + ventilation + chauffage + pilotage demontrant >= 30% gain",
     "Bonification importante (jusqu a x4)", "Selon volume",
     "kWh cumac bonifies x cours marche CEE",
     "Audit energetique prealable - simulation gain >= 30% en EF - travaux conformes",
     "Audit obligatoire par bureau d'etudes qualifie OPQIBI 1905",
     "Decret tertiaire (mise en conformite), Fonds Chaleur, regions",
     "HUARD + BE + mandataire", "Charte en vigueur",
     "Arrete coup de pouce renovation tertiaire",
     "https://www.ecologie.gouv.fr/coup-pouce-economies-denergie",
     "L'angle 'mise en conformite Decret Tertiaire + financement maximal' = pitch porte d'entree.",
     "Tres haute"),
]

# --- Photovoltaique / solaire ---
D += [
    ("Photovoltaique", "Prime a l'investissement / autoconsommation PV (S26 trimestriel)", "Prime nationale (arrete tarifaire)",
     "Etat (CRE / EDF OA)",
     "Toute personne morale ou physique - installations en autoconso avec vente de surplus, <= 500 kWc (au-dela : appel d'offres CRE)",
     "ELECTRICITE",
     "Installation PV en toiture ou ombriere, autoconsommation avec vente de surplus",
     "Variable par tranche - 0 pour > 100 kWc",
     "Tranche <= 3 kWc : ~80 EUR/kWc - 3-9 kWc : ~80 EUR/kWc - 9-36 kWc : ~190 EUR/kWc - 36-100 kWc : ~100 EUR/kWc (T2 2026 indicatif)",
     "EUR/kWc verse en 5 ans (1/5 par an) - revision trimestrielle par la CRE",
     "Installation en toiture/ombriere/hangar - integration paysagere - pose par installateur RGE - puissance <= 500 kWc",
     "QualiPV / QualibatPV / Qualifelec SPV (selon donneur d'ordre)",
     "TVA 10%, exoneration TVA depuis 2025 sur installations residentielles <= 9 kWc, aides regions",
     "EDF OA via portail dedie - HUARD pour montage technique",
     "Arrete tarifaire S26 (revision trimestrielle - avr/juil/oct/janv 2026)",
     "Arrete du 06/10/2021 modifie + arrete S26 2026",
     "https://www.photovoltaique.info/fr/preparer-un-projet/aspects-economiques/tarifs-dachat-et-autoconsommation/",
     "Marche tertiaire/industrie : autoconso = ROI 6-9 ans. Renforce avec obligation APER ombrieres parkings.",
     "Tres haute"),

    ("Photovoltaique", "Tarif d'achat surplus / totalite (Obligation d'Achat EDF OA)", "Tarif reglemente",
     "Etat / EDF OA",
     "Producteurs PV <= 500 kWc",
     "ELECTRICITE",
     "Rachat du surplus d'autoconsommation ou de la totalite de la production",
     "Contrat 20 ans", "N/A (tarif indexe)",
     "Tarif EUR/kWh fixe sur 20 ans - revision trimestrielle des nouveaux contrats",
     "Installation conforme - puissance <= 500 kWc - integration au reseau",
     "QualiPV / equivalent", "Prime autoconsommation",
     "EDF OA - dossier en ligne", "Contrat 20 ans, signature des MES",
     "Code de l'energie art. L.314-1 et suivants",
     "https://www.edf-oa.fr",
     "Securisation 20 ans du revenu = argument financier solide pour COMEX clients.",
     "Tres haute"),

    ("Photovoltaique", "Loi APER - obligation PV/ombrieres sur parkings et batiments", "Obligation reglementaire (creatrice de marche)",
     "Etat (DGEC)",
     "Proprietaires de parkings > 1500 m2 (depuis 07/2026 : 500 m2 pour les nouveaux), batiments commerciaux/industriels > 500 m2",
     "ELECTRICITE",
     "Couverture par ombrieres PV sur 50% de la surface du parking - integration PV en toiture sur batiments non residentiels neufs et renoves",
     "N/A - obligation",
     "Sanction jusqu'a 40000 EUR/an non couverte (parkings)",
     "Obligation - investissement client",
     "Calendrier d'application progressif - 07/2026 pour parkings 500-1500 m2",
     "QualiPV / Qualifelec SPV",
     "Prime autoconso, tarif rachat surplus, aides regions, suramortissement industrie",
     "Client (HUARD = installateur recommande)",
     "Loi 2023-175 + decrets 2023 + decret 2024 elargissement",
     "Loi APER + decrets 2023-1408 et 2024-XX",
     "https://www.legifrance.gouv.fr/jorf/id/JORFTEXT000047283081",
     "MARCHE CAPTIF MAJEUR : prospection systematique parkings hyper/super, zones logistiques, centres commerciaux, sites industriels.",
     "Tres haute"),

    ("Photovoltaique", "Programme OSCAR - ombrieres PV sur parkings publics", "Programme CEE (PRO-INNO-XX)",
     "Etat - finance via CEE",
     "Collectivites et acteurs publics",
     "ELECTRICITE",
     "Etudes et accompagnement a la pose d'ombrieres PV - financement partiel des etudes",
     "Plusieurs dizaines de milliers EUR / etude",
     "Couverture jusqu a 100% des etudes",
     "Forfait par etude", "Collectivite ou bailleur public - parking eligible",
     "BE qualifie", "Aides ADEME, prime autoconso PV",
     "Collectivite (HUARD = partenaire technique)",
     "Programme actif 2024-2027",
     "Arrete portant validation du programme OSCAR",
     "https://www.programme-oscar.fr",
     "Porte d'entree collectivites IDF - meme si l'aide va au client, HUARD entre comme installateur PV.",
     "Moyenne"),

    ("Photovoltaique", "Appels d'offres CRE PV (> 500 kWc)", "Appel d'offres",
     "CRE - Commission de regulation de l'energie",
     "Personnes morales - centrales PV au sol, ombrieres, batiments > 500 kWc",
     "ELECTRICITE",
     "Centrales PV de grande puissance - vente totale au reseau",
     "Tarif laureat sur 20 ans", "N/A",
     "Tarif EUR/MWh propose par le candidat - selection au moins-disant",
     "Surface eligible - foncier - permis - candidature dans la fenetre AO",
     "QualiPV / qualif. bureau etudes", "Cumul limite avec autres aides",
     "Client + AMOA", "AO PPE2 - sessions semestrielles",
     "Decisions CRE - PPE 2024-2035",
     "https://www.cre.fr/Pages-annexes/Recherche-Pages-annexes?query=appel+offres+PV",
     "Cible : sites industriels, logistique, foncier client > 5000 m2 toiture - HUARD partenaire d'installation.",
     "Moyenne"),
]

# --- ADEME / Fonds Chaleur / Decarbonation ---
D += [
    ("ADEME", "Fonds Chaleur ADEME - chaleur renouvelable", "Subvention",
     "ADEME",
     "Toute personne morale (entreprises, collectivites, copropriete, bailleurs sociaux, agricoles)",
     "CVC",
     "Geothermie (PAC sondes/nappe), biomasse, solaire thermique, reseaux de chaleur EnR&R, recuperation chaleur fatale, methanisation",
     "Variable selon taille - planchers de surface/puissance",
     "Jusqu'a 65% des couts eligibles HT (PME) - taux module selon taille entreprise et type",
     "Forfait EUR/MWh produit (mode forfaitaire jusqu a un seuil) ou aide a l'investissement (taux)",
     "Etude de faisabilite - production EnR significative - performance minimale",
     "Bureau d'etudes qualifie (OPQIBI 2008/2009/...)",
     "CEE, aides regions (sous conditions de non double subvention)",
     "Client (HUARD aide au montage)",
     "Enveloppe annuelle - appels a projets en continu (BCIAT, BCIB, etc.)",
     "Decisions ADEME annuelles - guide Fonds Chaleur",
     "https://fondschaleur.ademe.fr",
     "ARGUMENT MAJEUR CVC EnR : subvention massive + CEE + sortie d'energie fossile. Cibler chaufferie collective, hopitaux, ehpad, hotels, piscines.",
     "Tres haute"),

    ("ADEME", "Programme ACTEE - batiments publics", "Programme CEE (operateur FNCCR)",
     "FNCCR / CEE",
     "Collectivites locales (communes, EPCI, departements)",
     "ELECTRICITE / CVC / MAINTENANCE",
     "Etudes, AMOA, financement de l'economiste, des audits, des SI de suivi energie pour batiments publics (ecoles, mairies, gymnases)",
     "Quelques milliers EUR a centaines de milliers EUR",
     "Forfait par lot d'etudes / appareils de mesure",
     "Aide par lot", "Collectivite candidate dans un AAP ACTEE",
     "BE qualifie", "Fonds Chaleur, CEE, DSIL",
     "Collectivite (HUARD = installateur sur travaux qui suivent)",
     "Programme ACTEE+ 2024-2026",
     "Convention ADEME-FNCCR",
     "https://www.programme-actee.fr",
     "Entree commerciale collectivites IDF - on suit les groupements de commandes issus d'ACTEE.",
     "Moyenne"),

    ("ADEME", "Diag Decarbon'Action", "Subvention Bpifrance/ADEME",
     "Bpifrance + ADEME",
     "PME (< 500 salaries)",
     "ELECTRICITE / CVC",
     "Diagnostic GES + plan d'action de decarbonation accompagne par expert agree",
     "0 (pris en charge 80%)", "4000 EUR HT (cout total ~5000 EUR)",
     "Forfait", "PME, demande via plateforme Bpifrance",
     "Expert Bpifrance agree", "Cumul CEE, Fonds Chaleur, regions",
     "Client (Bpifrance)", "Programme en cours",
     "Bpifrance - Diag Decarbon'Action",
     "https://www.bpifrance.fr/catalogue-offres/transition-ecologique-et-energetique/diag-decarbonaction",
     "Porte d'entree IDEALE chez nos clients PME pour ensuite vendre les actions concretes (CVC, eclairage, PV).",
     "Haute"),

    ("ADEME", "Diag Eco-Flux", "Subvention Bpifrance/ADEME",
     "Bpifrance + ADEME",
     "PME 20-250 salaries multi-sites souvent",
     "ELECTRICITE / CVC / MAINTENANCE",
     "Accompagnement 12 mois reduction conso eau/energie/matieres",
     "0 (pris en charge ~75%)", "3750 EUR HT pour le PME",
     "Forfait", "PME industrielle/tertiaire/commerce",
     "Cabinet agree Bpifrance", "Cumul CEE, Fonds Chaleur",
     "Client (Bpifrance)", "Programme en cours",
     "Bpifrance Diag Eco-Flux",
     "https://www.bpifrance.fr/catalogue-offres/transition-ecologique-et-energetique/diag-eco-flux",
     "Etudes faites par expert = travaux ensuite confies a HUARD (qui aura aide a la conviction commerciale).",
     "Haute"),

    ("ADEME / France 2030", "Tremplin pour la transition ecologique des PME", "Subvention forfaitaire",
     "ADEME",
     "TPE/PME",
     "ELECTRICITE / CVC / MAINTENANCE",
     "Forfait sur ~50 actions eco (eclairage, pompes a chaleur, isolation, mobilite, etc.)",
     "1000 EUR (seuil minimal)", "200 000 EUR cumules par PME",
     "Forfait par action - taux 50 a 80%",
     "TPE/PME - depot en ligne ADEME - non commencee a la demande",
     "Aucune RGE specifique (selon action)",
     "Cumul CEE strict (l'aide se cumule mais decomptee)",
     "Client (ADEME)", "Guichet ouvert en continu",
     "Tremplin PME ADEME",
     "https://agirpourlatransition.ademe.fr/entreprises/dispositif-aide/tremplin-transition-ecologique-pme",
     "Outil terrain simple - couple parfait avec un devis HUARD eclairage/CVC pour TPE-PME.",
     "Haute"),

    ("ADEME / Industrie", "Decarbonation industrielle - BCIAT / IndusEE / Heat&Cool", "Subvention appel a projets",
     "ADEME / France 2030",
     "Industriels - sites de production",
     "CVC / MAINTENANCE",
     "Chaleur biomasse, electrification procedes, geothermie industrielle, recup chaleur fatale, efficacite procedes",
     "Variable", "Plusieurs M EUR par projet (regime aide d'Etat)",
     "Taux selon taille entreprise et zone AFR",
     "Audit prealable - business plan - candidature AAP",
     "BE qualifie", "CEE, Fonds Chaleur, regions",
     "Client (ADEME via plateforme)", "AAP annuels - BCIAT 2026 en cours",
     "Decisions ADEME / Plan France 2030",
     "https://entreprises.ademe.fr/dispositif-aide/decarbonation-industrie",
     "Cible : nos clients industriels IDF + grandes ETI - HUARD partenaire CVC/elec sur les projets.",
     "Moyenne"),
]

# --- Bpifrance & financements ---
D += [
    ("Financement", "Pret Vert Bpifrance", "Pret bonifie",
     "Bpifrance",
     "PME, ETI (> 3 ans)",
     "ELECTRICITE / CVC",
     "Investissements transition ecologique : equipements EnR, PAC, eclairage perf, batiments perf",
     "50 000 EUR", "Jusqu a 15 M EUR",
     "Pret 3-10 ans - taux preferentiel - sans garantie reelle ni caution personnelle",
     "Bilan positif - projet eligible TEE - ETI ou PME",
     "N/A", "Aides ADEME, CEE, regions",
     "Client (Bpifrance)", "Guichet ouvert",
     "Bpifrance - catalogue Pret Vert",
     "https://www.bpifrance.fr/catalogue-offres/transition-ecologique-et-energetique/pret-vert",
     "Levier financier pour gros chantiers HUARD - le client emprunte, on encaisse les travaux.",
     "Haute"),

    ("Financement", "Pret Economies d'Energie (PEE) Bpifrance/ADEME", "Pret bonifie",
     "Bpifrance + ADEME",
     "TPE/PME (> 3 ans)",
     "ELECTRICITE / CVC",
     "Equipements eligibles CEE (eclairage LED, PAC, regulation, isolation, ECS, vehicules elec)",
     "10 000 EUR", "500 000 EUR (et 3M EUR EnR)",
     "Pret 3-7 ans - taux fixe bonifie",
     "PME - factures detaillees - delai 12 mois", "N/A",
     "Cumul CEE et autres aides",
     "Client (Bpifrance)", "Guichet ouvert",
     "Bpifrance PEE",
     "https://www.bpifrance.fr/catalogue-offres/transition-ecologique-et-energetique/pret-economies-denergie",
     "Outil tres operationnel : on chiffre + on indique au client comment financer.",
     "Haute"),
]

# --- IRVE / mobilite electrique ---
D += [
    ("IRVE", "Programme ADVENIR - bornes recharge VE", "Programme CEE",
     "Avere-France / finance par CEE",
     "Copropriete, parkings entreprise (salaries+flottes), parkings ouverts au public, voirie, hotel/restauration, services autoroutiers",
     "ELECTRICITE",
     "Pose de bornes de recharge VE 3.7 a 350 kW, infrastructure collective copro, retrofit",
     "Variable selon categorie",
     "Plafonds 2026 : copro indiv. ~960 EUR - copro collective ~1660 EUR/point + 3000 EUR pre-equipement - parking entreprise salaries 1700 EUR - parking ouvert public 2700 EUR (standard) jusqu'a 15000 EUR (>140 kW)",
     "% du HT plafonne", "Devis signe APRES eligibilite ADVENIR - installateur labellise Advenir - bornes labellisees",
     "IRVE (qualifelec ou equivalent niveau IRVE-2 a 3)",
     "Aides regionales, suramortissement (entreprise)",
     "Installateur (HUARD) labellise ADVENIR",
     "Programme prolonge jusqu au 31/12/2027 (sous reserve enveloppe)",
     "Convention Avere-Etat - cahier des charges ADVENIR",
     "https://advenir.mobi",
     "Marche en explosion - HUARD doit etre installateur labellise ADVENIR. Pitch copro + entreprises avec flotte societe.",
     "Tres haute"),

    ("IRVE", "Decret tertiaire stationnement / Loi LOM - quotas bornes parkings non residentiels", "Obligation reglementaire",
     "Etat",
     "Proprietaires de batiments tertiaires neufs ou renoves",
     "ELECTRICITE",
     "Pre-equipement et points de recharge : seuils selon type batiment (au moins 5% places en pre-equipement neuf)",
     "N/A", "Sanctions reglementaires",
     "Obligation - investissement client",
     "Permis de construire deposes apres seuil - renovation lourde",
     "IRVE", "ADVENIR, aides regions",
     "Client", "En vigueur - renforcement 2025",
     "Code de la construction CCH art. R.111-14-2 et suivants + loi LOM",
     "https://www.legifrance.gouv.fr",
     "Argument obligation pour nos clients amenageurs / promoteurs / bailleurs - aide ADVENIR cumulable.",
     "Haute"),
]

# --- Decret tertiaire et obligations ---
D += [
    ("Obligation reglementaire", "Decret Tertiaire (DEET) - reduction conso - 40% / -50% / -60%",
     "Obligation - OPALE-MAR", "Etat (ADEME / OPERAT)",
     "Tous batiments / proprietaires / locataires tertiaire > 1000 m2",
     "ELECTRICITE / CVC / MAINTENANCE",
     "Travaux d'efficacite energetique (enveloppe, CVC, eclairage, GTB) pour atteindre les objectifs reglementaires",
     "N/A - obligation",
     "Sanctions : amende 1500 EUR particuliers, 7500 EUR personnes morales + name & shame",
     "Trajectoire valeur absolue (Cabs) ou en reduction relative (Crelat)",
     "Declaration annuelle sur OPERAT - objectif 2030 atteint",
     "Audit / accompagnement par BE", "TOUTES aides : CEE bonifies, ADEME, regions, Bpifrance",
     "Client (avec aide BE et HUARD)",
     "Echeances : -40% en 2030, -50% en 2040, -60% en 2050",
     "Decret 2019-771 + arrete methode 2020",
     "https://operat.ademe.fr",
     "PORTE D'ENTREE NUMERO 1 : creer un 'rendez-vous decret tertiaire' avec tous nos clients > 1000 m2. Audit gratuit -> plan d'actions -> chantiers HUARD + financement aides.",
     "Tres haute"),

    ("Obligation reglementaire", "Decret BACS - obligation GTB", "Obligation reglementaire",
     "Etat",
     "Tertiaire avec systeme CVC > 290 kW (existant : 01/01/2025) puis > 70 kW (01/01/2027)",
     "ELECTRICITE / CVC / MAINTENANCE",
     "Installation d'un systeme de GTB classe A ou B pilotant chauffage, refroidissement, ECS, eclairage",
     "N/A - obligation", "Sanctions reglementaires",
     "Obligation", "Echeance 01/01/2025 (>290 kW) - 01/01/2027 (>70 kW)",
     "Aucune RGE", "CEE BAT-TH-116 (bonifie), ADEME ACTEE",
     "Client", "Echeance imperative",
     "Decret n 2020-887 + arrete 02/06/2021",
     "https://www.ecologie.gouv.fr/decret-bacs",
     "Marche oblige - prospection systematique tous clients tertiaire avec puissance CVC concernee.",
     "Tres haute"),
]

# --- Aides regionales Ile-de-France ---
D += [
    ("Region IDF", "Aide regionale - efficacite energetique batiments des entreprises", "Subvention",
     "Region Ile-de-France",
     "PME et ETI franciliennes",
     "ELECTRICITE / CVC",
     "Audit, travaux d'efficacite, equipements EnR sur sites en IDF",
     "Variable selon AAP", "Plafond 200000 EUR (regime de minimis)",
     "Taux selon taille entreprise et type", "PME/ETI IDF - investissement en IDF",
     "BE qualifie", "Cumul CEE, ADEME, Bpifrance (regle de minimis)",
     "Client", "AAP regionaux annuels",
     "Deliberations CR IDF",
     "https://www.iledefrance.fr/aides-services/transition-energetique",
     "Argument 'aide regionale dans votre territoire' - facilite cloture du dossier.",
     "Moyenne"),

    ("Region IDF", "Energies POSIT'IF - tiers-financement copropriete", "Tiers-financement / Pret",
     "SEM Energies POSIT'IF (Region IDF)",
     "Coproprietes franciliennes",
     "CVC / ELECTRICITE",
     "Renovation energetique globale copro : enveloppe, CVC, ECS, eclairage commun",
     "Variable", "Selon projet (M EUR)",
     "Tiers-financement integral ou partiel",
     "Vote AG copro - audit", "BE qualifie / MOE",
     "Cumul MaPrimeRenov Copro, CEE, regions",
     "Syndic + Energies POSIT'IF", "Service permanent",
     "Statuts SEM Energies POSIT'IF",
     "https://www.energiespositif.fr",
     "Levier majeur copro IDF - HUARD se positionne comme entreprise de travaux dans les groupements.",
     "Haute"),

    ("Region IDF", "Plan SESAME / aides Region IDF developpement economique vert", "Subvention",
     "Region IDF / Bpifrance",
     "PME franciliennes",
     "ELECTRICITE / CVC / MAINTENANCE",
     "Investissements lies a la transition energetique des sites de production",
     "Variable", "Plafond regime aide d'Etat",
     "Aide en EUR - taux selon programme", "PME IDF - investissement en IDF",
     "Variable", "Cumul ADEME / Bpifrance / CEE",
     "Client", "AAP recurrents",
     "Deliberations CR IDF / convention Bpifrance",
     "https://www.iledefrance.fr",
     "Activer le partenariat Bpifrance HUARD pour orienter les clients.",
     "Moyenne"),

    ("Metropole / IDF", "Metropole du Grand Paris - Fonds d'investissement metropolitain", "Subvention",
     "Metropole du Grand Paris",
     "Communes et EPT du Grand Paris",
     "ELECTRICITE / CVC",
     "Equipements publics - efficacite energetique, EnR, eclairage public, IRVE",
     "Variable", "Plafonds par operation",
     "Subvention", "Projet sur territoire MGP",
     "Variable", "Cumul programme ACTEE, CEE",
     "Collectivite", "AAP annuels MGP",
     "Deliberations MGP",
     "https://www.metropolegrandparis.fr",
     "Cible directe : marches publics MGP - veille AAP a tenir mensuellement.",
     "Moyenne"),
]

# --- Aides residentielles / copropriete (utile car copros = clients HUARD) ---
D += [
    ("Residentiel / Copropriete", "MaPrimeRenov' Copropriete", "Prime",
     "Anah",
     "Coproprietes (syndicat des coproprietaires)",
     "ELECTRICITE / CVC",
     "Renovation globale : isolation, chauffage collectif, ventilation, ECS, eclairage commun, IRVE collective",
     "Gains 35% mini (geste copro)",
     "Jusqu a 25% du cout HT plafonne (geste +15% si gain >= 50%) + bonus sortie passoire / BBC",
     "% des couts HT plafonne par logement", "Audit prealable - vote AG - >= 75% lots residentiels",
     "Mandataire MAR (Mon Accompagnateur Renov) + RGE pour entreprises",
     "Cumul CEE, Eco-PTZ copro, regions, Energies POSIT'IF",
     "Syndic + mandataire", "Dispositif perenne - baremes 2026",
     "Decret MPR Copro 2024",
     "https://france-renov.gouv.fr/aides/mpr",
     "Marche copro = recurrent fort en IDF. HUARD se positionne sur chauffage collectif + eclairage parties communes + IRVE.",
     "Tres haute"),

    ("Residentiel", "MaPrimeRenov' - parcours par geste et accompagne", "Prime",
     "Anah",
     "Proprietaires occupants et bailleurs (logements > 15 ans)",
     "CVC / ELECTRICITE",
     "PAC, chaudieres biomasse, ventilation, isolation, audit - parcours accompagne avec MAR pour gains >= 35%",
     "Quelques centaines EUR (geste)", "Jusqu a 70000 EUR (renovation d'ampleur tres modeste)",
     "Forfait par geste OU % des couts en parcours accompagne (jusqu'a 90% modeste, 80% tres modeste)",
     "Logement > 15 ans, residence principale", "RGE obligatoire",
     "Cumul CEE, Eco-PTZ, regions, TVA 5,5%",
     "Particulier (avec MAR si accompagne)",
     "Dispositif perenne",
     "Decrets MPR annuels",
     "https://france-renov.gouv.fr",
     "Pour particuliers HUARD : remplacement chaudiere par PAC, audit, accompagnement complet.",
     "Haute"),

    ("Residentiel", "Eco-PTZ - Eco-pret a taux zero", "Pret a taux zero",
     "Etat (banques partenaires)",
     "Particuliers, coproprietes (eco-PTZ copro)",
     "ELECTRICITE / CVC",
     "Bouquet de travaux d'efficacite energetique ou geste unique RGE",
     "7000 EUR (geste simple)", "50000 EUR (renovation globale - depuis 2022)",
     "Pret amortissable jusqu a 20 ans - 0% interet",
     "Logement > 2 ans - residence principale - RGE", "RGE obligatoire",
     "Cumul MaPrimeRenov, CEE, regions",
     "Particulier / syndic via banque",
     "Prolonge jusque 31/12/2027",
     "Art. 244 quater U du CGI",
     "https://france-renov.gouv.fr/aides/eco-ptz",
     "Argument financement zero apport - lever objection 'pas de tresorerie'.",
     "Moyenne"),

    ("Residentiel", "TVA reduite 5,5% travaux renovation energetique", "TVA reduite",
     "Etat",
     "Particuliers, bailleurs, coprietes pour travaux dans logements > 2 ans",
     "ELECTRICITE / CVC",
     "Travaux d'amelioration energetique + travaux induits indissociables",
     "N/A", "N/A",
     "TVA 5,5% au lieu de 20% sur main d'oeuvre + materiel",
     "Logement > 2 ans, residence principale ou secondaire - attestation client",
     "RGE", "Cumul integral",
     "HUARD facture directement", "Permanent",
     "Art. 278-0 bis A du CGI",
     "https://www.economie.gouv.fr/particuliers/taux-tva-renovation-logement",
     "Reflexe systematique sur tous devis particuliers - 14,5 points de marge effective.",
     "Haute"),

    ("Residentiel", "TVA 10% - travaux amelioration / amenagement logement > 2 ans", "TVA reduite",
     "Etat",
     "Particuliers, bailleurs",
     "ELECTRICITE / CVC / MAINTENANCE",
     "Travaux non eligibles 5,5% mais d'amelioration / transformation / amenagement",
     "N/A", "N/A", "TVA 10%",
     "Logement > 2 ans - attestation client", "Aucune",
     "Cumul integral", "HUARD facture", "Permanent",
     "Art. 279-0 bis CGI",
     "https://www.economie.gouv.fr/particuliers/taux-tva-renovation-logement",
     "Reflexe systematique - bien distinguer 5,5% / 10% selon nature.",
     "Haute"),
]

# --- Programmes specifiques ---
D += [
    ("Programme", "PRO-SMEn - aide certification ISO 50001", "Subvention forfaitaire",
     "ATEE - finance par CEE",
     "Entreprises certifiees ISO 50001",
     "MAINTENANCE / CVC",
     "Aide au cout de certification du systeme de management de l'energie",
     "20% du CA energie n-1",
     "40000 EUR (PME) / 100000 EUR (au-dela)",
     "Forfait base sur facture energie",
     "Certification ISO 50001 obtenue - depot dans les 12 mois",
     "Auditeur ISO 50001", "Cumul autres aides",
     "Client (ATEE)", "Programme en cours",
     "Convention ATEE-Etat",
     "https://www.pro-smen.org",
     "Pour clients industriels - on les accompagne sur le volet metier energetique.",
     "Faible"),

    ("Programme", "PROFEEL - performance energetique batiment", "Programme CEE",
     "AQC + filiere - CEE",
     "Professionnels, MO public/prive",
     "ELECTRICITE / CVC",
     "Outils, methodes, formations sur performance batiment",
     "Outils gratuits", "N/A",
     "N/A - mise a disposition outils", "Acces ouvert",
     "N/A", "N/A", "N/A", "Programme en cours",
     "Convention AQC", "https://programmeprofeel.fr",
     "Utiliser les outils PROFEEL pour fiabiliser les devis et arguments techniques.",
     "Faible"),

    ("Programme", "Boost'Eco / dispositifs Conseillers en energie partages", "Programme CEE",
     "FNCCR / collectivites",
     "Petites communes, EPCI",
     "ELECTRICITE / CVC",
     "Conseiller en energie mutualise pour gerer le patrimoine batimentaire public",
     "Co-finance", "N/A",
     "Variable selon convention", "Convention avec syndicat departemental",
     "N/A", "Cumul ACTEE",
     "Collectivite", "Programme en cours",
     "Convention FNCCR-CEE", "https://www.fnccr.asso.fr",
     "Levier indirect : les conseillers prescrivent souvent les travaux HUARD.",
     "Faible"),
]

# Inserer les donnees
priority_color = {"Tres haute": GREEN, "Haute": "82C341", "Moyenne": "F4A261", "Faible": "BFC9D9"}
cat_colors = {
    "CEE - Tertiaire": "DCE7F3",
    "CEE - Industrie": "C7D9EC",
    "Coup de pouce CEE": "FFE9C4",
    "Photovoltaique": "FFF4B8",
    "ADEME": "D9F0E3",
    "ADEME / France 2030": "D9F0E3",
    "ADEME / Industrie": "D9F0E3",
    "Financement": "EAD9F0",
    "IRVE": "F9D5C2",
    "Obligation reglementaire": "F3C7C2",
    "Region IDF": "E0F0F0",
    "Metropole / IDF": "E0F0F0",
    "Residentiel": "F0E6CF",
    "Residentiel / Copropriete": "F0E6CF",
    "Programme": "EFEFEF",
}

for i, row in enumerate(D, start=1):
    excel_row = 3 + i
    # Numero
    ws2.cell(row=excel_row, column=1, value=i)
    # Categorie ... toutes les valeurs sauf la priorite
    for col_idx, val in enumerate(row, start=2):
        ws2.cell(row=excel_row, column=col_idx, value=val)
    # Style
    style_data_row(ws2, excel_row, len(headers), zebra=(i % 2 == 0))
    # Couleur categorie
    cat = row[0]
    if cat in cat_colors:
        ws2.cell(row=excel_row, column=2).fill = PatternFill("solid", fgColor=cat_colors[cat])
    # Priorite avec code couleur
    prio = row[-1]
    pc = ws2.cell(row=excel_row, column=20)
    pc.fill = PatternFill("solid", fgColor=priority_color.get(prio, "BFC9D9"))
    pc.font = Font(bold=True, color="FFFFFF" if prio in ("Tres haute", "Haute") else "1F2F4D", size=10)
    pc.alignment = Alignment(horizontal="center", vertical="center")
    # Lien hypertexte
    link_cell = ws2.cell(row=excel_row, column=18)
    if isinstance(link_cell.value, str) and link_cell.value.startswith("http"):
        link_cell.hyperlink = link_cell.value
        link_cell.font = Font(color=BLUE, underline="single", size=9)
    # Hauteur
    ws2.row_dimensions[excel_row].height = 95

# Largeurs des colonnes (en caracteres)
widths = [5, 16, 28, 14, 18, 22, 18, 36, 14, 28, 22, 30, 20, 22, 18, 18, 22, 30, 38, 12]
set_widths(ws2, widths)

# Filtre auto
ws2.auto_filter.ref = f"A3:T{3+len(D)}"

# =========================================================================
# 3. VUE PAR ACTIVITE HUARD
# =========================================================================
ws3 = wb.create_sheet("02 - Par activite HUARD")
ws3.sheet_view.showGridLines = False
title_cell(ws3, "A1", "DISPOSITIFS PAR ACTIVITE HUARD", size=16)
ws3.merge_cells("A1:F1")
ws3.row_dimensions[1].height = 30

activites = [
    ("ELECTRICITE (courants forts/faibles, eclairage, distribution, IRVE, PV)", [
        ("CEE BAT-EQ-127 - Luminaires LED tertiaire", "Tres haute", "CEE", "Toiture commerciale obligation APER pour parkings - relamping LED"),
        ("CEE BAT-EQ-133 - Gestion de l'eclairage (detection, gradation)", "Haute", "CEE", "Couple systematiquement avec BAT-EQ-127"),
        ("Prime autoconsommation PV S26 + Tarif EDF OA", "Tres haute", "Prime/tarif", "Marche tertiaire/industrie - autoconso 9-500 kWc"),
        ("Loi APER - ombrieres PV parkings > 1500 m2 (500 m2 en 07/2026)", "Tres haute", "Obligation", "Marche captif - prospecter parkings clients"),
        ("Programme ADVENIR - bornes recharge IRVE", "Tres haute", "Programme CEE", "Copro + entreprises - HUARD doit etre labellise ADVENIR"),
        ("CEE BAT-TH-116 - GTB + decret BACS", "Tres haute", "CEE + obligation", "Obligation > 290 kW (2025) puis > 70 kW (2027)"),
        ("CEE IND-UT-103 - Variateurs vitesse moteurs", "Haute", "CEE", "Sites industriels - moteurs charge variable"),
        ("Loi LOM - pre-equipement IRVE parkings tertiaires", "Haute", "Obligation", "Tous batiments tertiaires neufs/renoves"),
    ]),
    ("CVC (chauffage, ventilation, climatisation, ECS, plomberie)", [
        ("CEE BAT-TH-104 - PAC air/eau, eau/eau", "Tres haute", "CEE", "Copro chauffage fioul/gaz - sortie energie fossile bonifiee"),
        ("Coup de pouce 'Chauffage tertiaire'", "Tres haute", "CEE bonifie", "Bonification x2-4 sortie fioul/charbon"),
        ("Fonds Chaleur ADEME - geothermie, biomasse, solaire thermique, chaleur fatale", "Tres haute", "Subvention", "Aide jusqu'a 65% - chaufferies collectives, hopitaux, hotels, piscines"),
        ("Coup de pouce 'Renovation performante tertiaire'", "Tres haute", "CEE bonifie", "Audit + bouquet >= 30% gain"),
        ("CEE BAT-TH-127 - Raccordement reseau de chaleur EnR&R", "Haute", "CEE", "100+ reseaux de chaleur en IDF"),
        ("CEE BAT-TH-146 - Calorifugeage points singuliers", "Moyenne", "CEE", "Action recurrente - integrer aux contrats P2/P3"),
        ("CEE BAT-TH-155/125 - VMC double flux", "Moyenne", "CEE", "QAI + economies"),
        ("MaPrimeRenov Copropriete", "Tres haute", "Prime", "Chauffage collectif copro IDF"),
    ]),
    ("TELECOMS & INFORMATIQUE (reseau cuivre/fibre, courants faibles, datacenters)", [
        ("Bonification CEE actions datacenter (efficacite refroidissement, free cooling)", "Moyenne", "CEE specifique", "Sur fiches IND-UT pour groupes froid datacenter"),
        ("CEE BAT-TH-116 - GTB pour pilotage CVC datacenter", "Haute", "CEE", "Indirect mais structurant pour clients hebergeurs"),
        ("Diag Decarbon'Action (Bpifrance)", "Haute", "Subvention", "Forfait 4000 EUR PME - utile sur clients telecoms/IT"),
        ("Aides regionales IDF transition numerique vertueuse", "Moyenne", "Subvention", "Selon AAP en cours"),
    ]),
    ("MAINTENANCE multi-technique (contrats P2/P3, GMAO, maintenance preventive)", [
        ("CEE BAT-TH-146 - Calorifugeage", "Haute", "CEE", "Integrer dans toute revisite annuelle chaufferie"),
        ("CEE BAT-TH-116 - GTB pilotage", "Tres haute", "CEE + obligation BACS", "Contrats d'exploitation incluant pilotage"),
        ("CEE BAT-TH-141 - Programmation horaire", "Moyenne", "CEE", "Action low-cost a integrer systematiquement"),
        ("CEE IND-UT-103 - Variateurs", "Moyenne", "CEE", "Equiper le parc moteurs"),
        ("Coup de pouce 'Pilotage connecte du chauffage'", "Haute", "CEE bonifie", "Argument 'maintenance connectee' HUARD"),
        ("PRO-SMEn - ISO 50001", "Faible", "Subvention", "Pour gros clients industriels qui certifient"),
    ]),
]

row = 3
for activite_titre, dispositifs in activites:
    ws3.cell(row=row, column=1, value=activite_titre).font = Font(size=14, bold=True, color="FFFFFF")
    ws3.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws3.cell(row=row, column=1).alignment = Alignment(vertical="center", indent=1)
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws3.row_dimensions[row].height = 28
    row += 1
    header_row(ws3, row, ["Dispositif", "Priorite", "Type", "Angle commercial / Cible", ""])
    ws3.row_dimensions[row].height = 24
    row += 1
    for d in dispositifs:
        for ci, v in enumerate(d, start=1):
            cell = ws3.cell(row=row, column=ci, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(size=10)
            if ci == 2 and v in priority_color:
                cell.fill = PatternFill("solid", fgColor=priority_color[v])
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True, color="FFFFFF" if v in ("Tres haute", "Haute") else "1F2F4D", size=10)
        ws3.row_dimensions[row].height = 36
        row += 1
    row += 1

set_widths(ws3, [42, 14, 18, 55, 8])

# =========================================================================
# 4. ELIGIBILITE CLIENTS - grille rapide
# =========================================================================
ws4 = wb.create_sheet("03 - Eligibilite clients")
ws4.sheet_view.showGridLines = False
title_cell(ws4, "A1", "GRILLE DE QUALIFICATION RAPIDE - QUI BENEFICIE DE QUOI", size=16)
ws4.merge_cells("A1:H1")
ws4.row_dimensions[1].height = 30

ws4["A2"] = "Cocher mentalement le profil client, lire les dispositifs eligibles sur la ligne."
ws4["A2"].font = Font(italic=True, color=GREY)
ws4.merge_cells("A2:H2")

grille = [
    ["Profil client", "CEE tertiaire", "Coup de pouce", "Fonds Chaleur", "Bpifrance Pret/Diag", "Region IDF", "ADVENIR IRVE", "MaPrimeRenov / Eco-PTZ"],
    ["PME tertiaire (bureaux, commerces, services) < 1000 m2", "OUI", "OUI (selon)", "OUI si EnR", "OUI", "OUI", "OUI parking entreprise", "NON"],
    ["Tertiaire 1000 - 5000 m2 (decret DEET)", "OUI", "OUI", "OUI EnR / RDC", "OUI", "OUI", "OUI", "NON"],
    ["Grand tertiaire > 5000 m2 (groupes, sieges)", "OUI", "OUI (selon)", "OUI EnR", "Pret Vert", "OUI (groupements)", "OUI", "NON"],
    ["Industrie / sites de production", "OUI (fiches IND)", "Limite", "OUI chaleur fatale, biomasse", "OUI", "OUI", "OUI flottes salaries", "NON"],
    ["Collectivite (commune, EPCI, MGP)", "OUI BAT", "OUI selon", "OUI", "Banque des territoires", "OUI MGP/IDF + ACTEE", "OUI voirie + bat. publics", "NON"],
    ["Bailleur social / OPH", "OUI BAR & BAT", "OUI", "OUI selon", "OUI", "OUI Energies POSIT'IF", "OUI parc", "MPR Copro pour mixte"],
    ["Copropriete (residence collective)", "OUI BAR-TH", "OUI", "OUI EnR collectif", "Eco-PTZ copro", "OUI POSIT'IF", "OUI infra collective", "OUI MPR Copro + Eco-PTZ"],
    ["Promoteur / Constructeur (neuf)", "Limite (rehab)", "Limite", "Selon", "OUI", "OUI", "OUI pre-equipement", "NON"],
    ["Particulier maison individuelle", "OUI BAR-TH/EN", "OUI", "Limite", "NON", "Selon dispo", "Particulier <= 960 EUR", "OUI MPR + Eco-PTZ"],
    ["Hotel / restauration / sante", "OUI tertiaire", "OUI", "OUI EnR (ECS solaire, biomasse)", "OUI", "OUI", "OUI", "NON"],
    ["Datacenter / IT", "Limite (fiches IND-UT)", "NON", "OUI chaleur fatale", "OUI", "OUI selon AAP", "OUI", "NON"],
    ["Logistique / entrepots", "OUI BAT", "OUI", "OUI selon", "OUI", "OUI", "OUI parking et flotte", "NON"],
]

for ri, line in enumerate(grille, start=4):
    for ci, val in enumerate(line, start=1):
        c = ws4.cell(row=ri, column=ci, value=val)
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center", wrap_text=True)
        c.border = border
        if ri == 4:
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=BLUE)
        else:
            c.font = Font(size=10)
            if ci == 1:
                c.font = Font(bold=True, color=NAVY, size=10)
            if val == "OUI":
                c.fill = PatternFill("solid", fgColor="D9F0E3")
            elif val == "NON":
                c.fill = PatternFill("solid", fgColor="F3C7C2")
            elif val == "Limite":
                c.fill = PatternFill("solid", fgColor="FFE9C4")
    ws4.row_dimensions[ri].height = 30

ws4.row_dimensions[4].height = 36
set_widths(ws4, [40, 16, 14, 18, 18, 16, 18, 22])

# =========================================================================
# 5. ARGUMENTAIRE COMMERCIAL
# =========================================================================
ws5 = wb.create_sheet("04 - Argumentaire")
ws5.sheet_view.showGridLines = False
title_cell(ws5, "A1", "ARGUMENTAIRE COMMERCIAL ET LEVIERS DE NEGOCIATION", size=16)
ws5.merge_cells("A1:D1")
ws5.row_dimensions[1].height = 30

argu = [
    ("Pitch d'entree (porte ouverte client)",
     "« Saviez-vous que sur vos travaux d'eclairage / chauffage / IRVE, l'Etat finance entre 25% et 70% via les CEE et les primes ADEME ? Nous avons cartographie 37 dispositifs - je viens vous montrer ceux qui s'appliquent a votre site. »",
     "Tertiaire, copropriete, industrie",
     "Tres haute"),
    ("Angle Decret Tertiaire (DEET)",
     "« Vous etes oblige par le decret tertiaire d'atteindre -40% de conso en 2030. Si vous ne deposez pas la declaration OPERAT, c'est 7500 EUR d'amende + name and shame. Nous proposons un parcours : audit -> plan -> travaux finances jusqu a 70% par CEE bonifies + ADEME. »",
     "Proprietaires/exploitants tertiaire > 1000 m2",
     "Tres haute"),
    ("Angle Decret BACS (GTB obligatoire)",
     "« Depuis le 1er janvier 2025, votre batiment doit etre equipe d'une GTB classe A/B si systeme CVC > 290 kW. Au 1er janvier 2027, le seuil descend a 70 kW. La fiche CEE BAT-TH-116 finance 50-80% de l'investissement. »",
     "Tertiaire avec CVC moyenne/grosse puissance",
     "Tres haute"),
    ("Angle Loi APER (ombrieres PV)",
     "« A partir de juillet 2026, tout parking de plus de 500 m2 devra etre couvert a 50% d'ombrieres PV. Pour un parking 1000 m2, c'est 80-100 kWc PV qui peut generer 13-15k EUR/an de revente surplus + prime autoconso. Nous montons le projet cle en main. »",
     "Hypermarches, logistique, zones commerciales, industries",
     "Tres haute"),
    ("Angle IRVE/copropriete",
     "« Vos coproprietaires demandent des bornes ? Le programme ADVENIR finance jusqu'a 50% de l'installation (1660 EUR par point + 3000 EUR pre-equipement). Nous etes installateur labellise ADVENIR + Qualifelec IRVE. »",
     "Coproprietes, entreprises avec flotte salaries",
     "Tres haute"),
    ("Angle 'zero cash' particulier",
     "« Pour le remplacement de votre chaudiere, vous beneficiez de : MaPrimeRenov + CEE + Eco-PTZ + TVA 5,5%. Pour un menage modeste, le reste a charge est souvent < 10% du devis. »",
     "Particuliers menage modeste a intermediaire",
     "Haute"),
    ("Angle 'efficacite + amorti' PME industrielle",
     "« Diag Eco-Flux Bpifrance (3 750 EUR pour vous, le reste finance) + CEE sur l'eclairage et les variateurs + Pret Economies d'Energie a taux bonifie : amortissement < 4 ans, marge brute amelioree de 2-3 points. »",
     "PME industrielles 20-250 salaries",
     "Haute"),
    ("Cumul possible (a maitriser)",
     "Regle d'or : CEE + Fonds Chaleur cumulables sous condition (eviter double subvention sur meme equipement). CEE + MPR Copro cumulables. ADVENIR + CEE = pas de cumul direct (ADVENIR est finance par CEE). TVA 5,5% cumul integral. Pret Vert / Eco-PTZ : cumul integral avec primes (financement du reste a charge).",
     "TRANSVERSAL",
     "Reference"),
    ("Reflexes facturation",
     "1) Toujours verifier RGE en cours de validite avant signature client residentiel. 2) Engagement CEE AVANT signature devis sinon perdu. 3) Attestation TVA 5,5% systematique. 4) Mentionner clairement 'travaux eligibles CEE / ADVENIR / Fonds Chaleur' dans le devis. 5) Mandataire CEE choisi en amont (ne pas faire des CEE en interne sauf si volume justifie).",
     "TRANSVERSAL",
     "Reference"),
    ("Objection 'C'est trop complique'",
     "« Justement - nous gerons les dossiers d'aide pour vous. Notre partenaire CEE/Mandataire monte les dossiers, on fait l'avance, vous touchez la prime. Pour Fonds Chaleur ou ADVENIR, on vous accompagne dans le depot. Cela fait partie de notre offre commerciale. »",
     "TOUS",
     "Reference"),
    ("Objection 'On a deja un installateur'",
     "« Sur les aides, nous reconnaissons que ce n'est pas un sujet de prix mais d'expertise du financement. Un audit gratuit du potentiel d'aides sur vos batiments peut faire emerger 50-150k EUR de financement non capte. Pas d'engagement. »",
     "TOUS",
     "Reference"),
]

header_row(ws5, 3, ["Cas commercial", "Pitch / explication", "Cible", "Priorite"])
ws5.row_dimensions[3].height = 32
for i, (cas, pitch, cible, prio) in enumerate(argu, start=4):
    ws5.cell(row=i, column=1, value=cas).font = Font(bold=True, color=NAVY, size=10)
    ws5.cell(row=i, column=2, value=pitch).font = Font(size=10)
    ws5.cell(row=i, column=3, value=cible).font = Font(italic=True, size=10)
    ws5.cell(row=i, column=4, value=prio)
    for col in range(1, 5):
        c = ws5.cell(row=i, column=col)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        if (i - 4) % 2 == 0:
            if c.fill.fgColor.rgb in (None, "00000000"):
                c.fill = PatternFill("solid", fgColor=LIGHT)
    if prio in priority_color:
        pc = ws5.cell(row=i, column=4)
        pc.fill = PatternFill("solid", fgColor=priority_color[prio])
        pc.font = Font(bold=True, color="FFFFFF" if prio in ("Tres haute", "Haute") else "1F2F4D")
        pc.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[i].height = 90

set_widths(ws5, [28, 90, 30, 14])

# =========================================================================
# 6. CEE FICHES CLES (vue technique)
# =========================================================================
ws6 = wb.create_sheet("05 - CEE fiches cles")
ws6.sheet_view.showGridLines = False
title_cell(ws6, "A1", "FICHES CEE STANDARDISEES - DETAILS TECHNIQUES", size=16)
ws6.merge_cells("A1:E1")
ws6.row_dimensions[1].height = 30

ws6["A2"] = "Liste resserree des fiches CEE les plus utilisees pour les chantiers HUARD. Source officielle : ecologie.gouv.fr"
ws6["A2"].font = Font(italic=True, color=GREY)
ws6.merge_cells("A2:E2")

cee_data = [
    ("Fiche", "Intitule", "Secteur", "Usage type HUARD", "Notes / pieges"),
    ("BAT-EQ-127", "Luminaire a modules LED (tertiaire)", "Tertiaire", "Relamping bureaux, commerces, parkings couverts", "Verifier efficacite >= 130 lm/W et systeme de gestion (sinon BAT-EQ-117/133 a coupler)"),
    ("BAT-EQ-133", "Systeme de gestion de l'eclairage", "Tertiaire", "DALI/KNX, detecteurs presence", "A coupler avec BAT-EQ-127 pour effet de levier max"),
    ("BAT-EQ-117", "Variation puissance eclairage exterieur", "Tertiaire", "Parkings exterieurs, eclairage site", "Variation par paliers ou horloge - asservissement effectif"),
    ("BAT-TH-104", "Pompe a chaleur (PAC) air/eau, eau/eau", "Tertiaire", "Remplacement chaudieres fioul/gaz copro et tertiaire", "SCOP minimal selon t emperature - bonifie sortie fioul"),
    ("BAT-TH-113", "PAC a absorption gaz", "Tertiaire", "Sites avec ressource gaz", "Moins courant - verifier rentabilite"),
    ("BAT-TH-116", "GTB (Gestion Technique du Batiment) classe A/B", "Tertiaire", "Pilotage CVC + eclairage + ECS - decret BACS", "Classe A/B obligatoire - certification de la solution"),
    ("BAT-TH-125", "Ventilation double flux haute efficacite", "Tertiaire", "Bureaux, ecoles, ehpad", "Echangeur >= 75% - filtre F7 - moteur ECM"),
    ("BAT-TH-127", "Raccordement reseau de chaleur EnR&R > 50%", "Tertiaire", "Copro et tertiaire raccordable", "Verifier taux EnR du reseau (declaration ADEME)"),
    ("BAT-TH-141", "Programmation/regulation chauffage", "Tertiaire", "Reglage horloge / programmateur", "Action complementaire systematique"),
    ("BAT-TH-145", "Isolation reseau (calorifugeage)", "Tertiaire", "Reseaux ECS et chauffage chaufferie", "Classes d'isolation - linaire de reseau"),
    ("BAT-TH-146", "Calorifugeage points singuliers", "Tertiaire", "Vannes, brides, pompes chaufferie", "Action recurrente facile a vendre en contrats P2/P3"),
    ("BAT-EN-101", "Isolation combles/toitures tertiaires", "Tertiaire", "Bati existant tertiaire", "R minimum selon zone"),
    ("BAT-EN-103", "Isolation murs", "Tertiaire", "ITE ou ITI", "R minimum, performance attestee"),
    ("IND-UT-103", "Variateur electronique de vitesse (VEV)", "Industrie", "Sites industriels - moteurs pompes/ventilo/compresseurs", "Charge variable indispensable - tres bon ROI"),
    ("IND-UT-117", "Recuperation chaleur compresseur air", "Industrie", "Sites avec air comprime continu", "Cumul Fonds Chaleur sur chaleur fatale"),
    ("IND-UT-102", "Recuperation chaleur groupe froid", "Industrie", "Industries agroalimentaires, datacenter", "Verifier reutilisation effective"),
    ("IND-UT-121", "Matelas isolant points singuliers industriels", "Industrie", "Reseaux vapeur, chaud", "Demontable - economique"),
    ("BAR-TH-104", "PAC air/eau, eau/eau (logements collectifs et individuels)", "Residentiel", "Maisons, copros", "RGE obligatoire - bonifications coup de pouce"),
    ("BAR-EN-101", "Isolation combles residentiel", "Residentiel", "Maisons individuelles, dernier etage copro", "Tres utilisee, baremes coup de pouce"),
]

header_row(ws6, 4, cee_data[0], fill=BLUE)
ws6.row_dimensions[4].height = 30
for i, row in enumerate(cee_data[1:], start=5):
    for ci, val in enumerate(row, start=1):
        c = ws6.cell(row=i, column=ci, value=val)
        c.border = border
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if ci == 1:
            c.font = Font(bold=True, color=NAVY, size=10)
        if (i - 5) % 2 == 0:
            c.fill = PatternFill("solid", fgColor=LIGHT)
    ws6.row_dimensions[i].height = 35

set_widths(ws6, [14, 38, 14, 40, 45])

# =========================================================================
# 7. DECRET TERTIAIRE (focus)
# =========================================================================
ws7 = wb.create_sheet("06 - Decret tertiaire")
ws7.sheet_view.showGridLines = False
title_cell(ws7, "A1", "DECRET TERTIAIRE (DEET) - OPPORTUNITE COMMERCIALE N 1", size=16)
ws7.merge_cells("A1:D1")
ws7.row_dimensions[1].height = 30

ws7["A2"] = "Source : Decret n 2019-771 du 23 juillet 2019. Plateforme : OPERAT (ADEME)."
ws7["A2"].font = Font(italic=True, color=GREY)
ws7.merge_cells("A2:D2")

decret_lines = [
    ("Qui est concerne ?", "Tous batiments / parties de batiment a usage tertiaire de surface plancher >= 1000 m2 - en France metropolitaine et outre-mer.", ""),
    ("Quels usages ?", "Bureaux, commerces, hotels, restaurants, sante, enseignement, sport, justice, administration, logistique, services, culture, salles d'audience.", ""),
    ("Objectif", "Reduction conso d'energie finale : -40% en 2030, -50% en 2040, -60% en 2050. Reference : annee comprise entre 2010 et 2019.", ""),
    ("Methode Crelat", "Reduction relative par rapport a l'annee de reference choisie.", ""),
    ("Methode Cabs", "Valeur absolue exprimee en kWh/m2/an, etablie par arrete (annexe specifique a chaque categorie d'activite).", ""),
    ("Declaration annuelle", "Plateforme OPERAT (ADEME). Echeance : 30 septembre N+1 pour declarer les consos de l'annee N. Premiere echeance ferme 30/09/2022 pour donnees 2020+2021+annee de reference.", ""),
    ("Sanctions", "Mise en demeure - amende administrative 1500 EUR (particulier), 7500 EUR (personne morale) - publication des contrevenants ('name and shame').", ""),
    ("Modulation des objectifs", "Possible (dossier technique) si contraintes : architecture, patrimoine, couts disproportionnes, changement d'activite.", ""),
    ("Levier commercial HUARD", "1) Audit gratuit OPERAT du client. 2) Identifier le delta vs objectif 2030. 3) Construire un plan d'actions 'CVC + eclairage + GTB + EnR' chiffre. 4) Trouver le financement (CEE bonifies + Coup de pouce tertiaire + Fonds Chaleur + Region IDF). 5) Realiser les travaux. 6) Mesure & suivi via GTB (lien avec contrat maintenance HUARD).", ""),
    ("Cible prioritaire", "Bureaux et commerces > 1000 m2 en IDF qui n'ont pas encore declare. Hopitaux, ehpad, hotels, ecoles, centres commerciaux, logistique. Verifier sur OPERAT le statut declaratif (donnees publiques).", ""),
]

header_row(ws7, 4, ["Sujet", "Description", "Notes"], fill=NAVY)
for i, (sujet, desc, notes) in enumerate(decret_lines, start=5):
    ws7.cell(row=i, column=1, value=sujet).font = Font(bold=True, color=NAVY, size=11)
    ws7.cell(row=i, column=2, value=desc).font = Font(size=10)
    ws7.cell(row=i, column=3, value=notes).font = Font(size=10)
    for col in range(1, 4):
        c = ws7.cell(row=i, column=col)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
    ws7.row_dimensions[i].height = 55

set_widths(ws7, [25, 90, 30])

# =========================================================================
# 8. AIDES IDF
# =========================================================================
ws8 = wb.create_sheet("07 - Aides IDF")
ws8.sheet_view.showGridLines = False
title_cell(ws8, "A1", "DISPOSITIFS ILE-DE-FRANCE - SPECIFIQUE TERRITOIRE", size=16)
ws8.merge_cells("A1:E1")
ws8.row_dimensions[1].height = 30

idf_aides = [
    ("Dispositif", "Operateur", "Cible", "Aide", "Lien"),
    ("Aide Region IDF - efficacite energetique entreprises", "Region IDF", "PME/ETI", "Subvention plafonnee (regime de minimis 200000 EUR/3 ans)", "https://www.iledefrance.fr"),
    ("Energies POSIT'IF", "SEM IDF", "Coproprietes IDF", "Tiers-financement renovation globale + AMOA", "https://www.energiespositif.fr"),
    ("Fonds d'investissement metropolitain (FIM)", "Metropole du Grand Paris", "Communes/EPT MGP", "Subvention equipements publics", "https://www.metropolegrandparis.fr"),
    ("Aides Departement 91 / 78 / 92 / 94", "CD91, CD78, CD92, CD94", "Communes, copros, particuliers (selon CD)", "Subventions ponctuelles - varier selon AAP", "Site de chaque CD"),
    ("Plan SESAME / aides developpement economique vert", "Region IDF + Bpifrance", "PME industrielles IDF", "Pret bonifie + subvention", "https://www.iledefrance.fr"),
    ("ADEME Direction regionale IDF - appels a projets territoriaux", "ADEME IDF", "Entreprises / Collectivites IDF", "Subvention via AAP regionaux Fonds Chaleur, decarbonation", "https://ile-de-france.ademe.fr"),
    ("ARENE Ile-de-France (Institut Paris Region)", "Institut Paris Region", "Collectivites + acteurs IDF", "Accompagnement, etudes, conferences", "https://www.institutparisregion.fr"),
    ("Region IDF - Aide Aero-thermique parc social", "Region IDF", "Bailleurs sociaux IDF", "Subvention sur PAC parc social", "https://www.iledefrance.fr"),
    ("AIRPARIF / aides QAI tertiaire", "Region IDF / ARS", "Tertiaire IDF", "Diagnostic et aide investissement ventilation", "https://www.airparif.asso.fr"),
]

header_row(ws8, 3, idf_aides[0], fill=NAVY)
ws8.row_dimensions[3].height = 28
for i, row in enumerate(idf_aides[1:], start=4):
    for ci, val in enumerate(row, start=1):
        c = ws8.cell(row=i, column=ci, value=val)
        c.border = border
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if (i - 4) % 2 == 0:
            c.fill = PatternFill("solid", fgColor=LIGHT)
        if ci == 5 and isinstance(val, str) and val.startswith("http"):
            c.hyperlink = val
            c.font = Font(color=BLUE, underline="single", size=9)
    ws8.row_dimensions[i].height = 38

set_widths(ws8, [40, 22, 30, 45, 40])

# =========================================================================
# 9. SOURCES & LIENS
# =========================================================================
ws9 = wb.create_sheet("08 - Sources & liens")
ws9.sheet_view.showGridLines = False
title_cell(ws9, "A1", "SOURCES OFFICIELLES ET CONTACTS UTILES", size=16)
ws9.merge_cells("A1:C1")
ws9.row_dimensions[1].height = 30

sources = [
    ("Theme", "Reference / Organisme", "Lien"),
    ("CEE - Fiches d'operations standardisees", "Ministere Transition Ecologique", "https://www.ecologie.gouv.fr/operations-standardisees-deconomies-denergie"),
    ("CEE - Coups de pouce", "Ministere Transition Ecologique", "https://www.ecologie.gouv.fr/coup-pouce-economies-denergie"),
    ("Decret Tertiaire (DEET) - OPERAT", "ADEME", "https://operat.ademe.fr"),
    ("Decret BACS", "Ministere Transition Ecologique", "https://www.ecologie.gouv.fr/decret-bacs"),
    ("Loi APER (PV ombrieres / batiments)", "Legifrance", "https://www.legifrance.gouv.fr/jorf/id/JORFTEXT000047283081"),
    ("Tarifs achat PV - arretes tarifaires", "Photovoltaique.info", "https://www.photovoltaique.info/fr/preparer-un-projet/aspects-economiques/tarifs-dachat-et-autoconsommation/"),
    ("EDF Obligation d'Achat (solaire)", "EDF OA", "https://www.edf-oa.fr"),
    ("Fonds Chaleur ADEME", "ADEME", "https://fondschaleur.ademe.fr"),
    ("Programmes ADEME (Tremplin PME, Diag)", "ADEME / Bpifrance", "https://agirpourlatransition.ademe.fr"),
    ("Bpifrance - catalogue offres TEE", "Bpifrance", "https://www.bpifrance.fr/catalogue-offres/transition-ecologique-et-energetique"),
    ("Programme ADVENIR (IRVE)", "Avere-France", "https://advenir.mobi"),
    ("Programme ACTEE (collectivites)", "FNCCR", "https://www.programme-actee.fr"),
    ("Programme OSCAR (ombrieres collectivites)", "Programme OSCAR", "https://www.programme-oscar.fr"),
    ("PRO-SMEn (ISO 50001)", "ATEE", "https://www.pro-smen.org"),
    ("MaPrimeRenov / France Renov", "Anah / France Renov", "https://france-renov.gouv.fr"),
    ("Region Ile-de-France - aides transition", "Region IDF", "https://www.iledefrance.fr/aides-services/transition-energetique"),
    ("Energies POSIT'IF (copro IDF)", "SEM IDF", "https://www.energiespositif.fr"),
    ("Metropole du Grand Paris", "MGP", "https://www.metropolegrandparis.fr"),
    ("ADEME Ile-de-France", "ADEME", "https://ile-de-france.ademe.fr"),
    ("Institut Paris Region (ex-ARENE)", "IPR", "https://www.institutparisregion.fr"),
    ("Photovoltaique info (ressource technique)", "Hespul", "https://www.photovoltaique.info"),
    ("AQC - Agence Qualite Construction", "AQC", "https://qualiteconstruction.com"),
    ("Annuaire RGE / Qualibat", "Qualibat", "https://www.qualibat.com"),
    ("Annuaire Qualifelec (electricite, IRVE, PV)", "Qualifelec", "https://www.qualifelec.fr"),
    ("Annuaire QualiPV", "Qualit'EnR", "https://www.qualit-enr.org"),
]

header_row(ws9, 3, sources[0], fill=NAVY)
ws9.row_dimensions[3].height = 26
for i, row in enumerate(sources[1:], start=4):
    for ci, val in enumerate(row, start=1):
        c = ws9.cell(row=i, column=ci, value=val)
        c.border = border
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if (i - 4) % 2 == 0:
            c.fill = PatternFill("solid", fgColor=LIGHT)
        if ci == 3 and isinstance(val, str) and val.startswith("http"):
            c.hyperlink = val
            c.font = Font(color=BLUE, underline="single", size=10)
    ws9.row_dimensions[i].height = 24

set_widths(ws9, [42, 38, 70])

# =========================================================================
# 10. PLAN ADMINISTRATIF - PIPELINE 6 PHASES
# =========================================================================
ws10 = wb.create_sheet("09 - Plan administratif")
ws10.sheet_view.showGridLines = False
title_cell(ws10, "A1", "PLAN ADMINISTRATIF - PROCESS HUARD DE MONTAGE DES AIDES CLIENT", size=16)
ws10.merge_cells("A1:G1")
ws10.row_dimensions[1].height = 30

ws10["A2"] = "Pipeline en 6 phases - de la detection a l'encaissement de l'aide par le client. RACI et delais cibles."
ws10["A2"].font = Font(italic=True, color=GREY, size=10)
ws10.merge_cells("A2:G2")

phases_headers = ["Phase", "Delai cible", "Responsable interne", "Acteurs externes", "Actions cles", "Pieces a produire", "Point critique"]
header_row(ws10, 4, phases_headers, fill=NAVY)
ws10.row_dimensions[4].height = 36

phases = [
    ("PHASE 1\nDETECTION & QUALIFICATION", "J0 - J+2",
     "Commercial terrain",
     "Client (decideur / DAF / syndic / responsable patrimoine)",
     "1. Visite ou RDV de qualification\n2. Identification du potentiel d'aides via la grille 'Eligibilite clients'\n3. Consultation OPERAT si tertiaire > 1000 m2\n4. Verification : batiment > 2 ans, RGE applicable, profil client",
     "Fiche d'opportunite client (template HUARD)\nGrille de qualification d'eligibilite\nNote OPERAT si DEET",
     "Ne JAMAIS signer de devis avant engagement CEE - sinon prime perdue."),

    ("PHASE 2\nPRE-ETUDE & CHIFFRAGE", "J+3 - J+10",
     "Charge d'affaires + Referent aides HUARD",
     "Mandataire CEE partenaire\nBE thermique si audit",
     "1. Visite technique chiffrage\n2. Calcul des aides : CEE (simulateur partenaire), Coup de pouce, Fonds Chaleur, ADVENIR, MPR Copro, Region\n3. Construction note d'opportunite chiffree (cout HT - aides - reste a charge - ROI)\n4. Choix du dispositif optimal (cumul vs exclusion)",
     "Note d'opportunite chiffree\nDevis pre-detaille\nSimulation CEE / ADVENIR / Fonds Chaleur",
     "Verifier le cumul : ADVENIR ne se cumule pas avec CEE - choisir la voie. Bonification Coup de pouce a engager avant fin de charte."),

    ("PHASE 3\nCONSTITUTION DU DOSSIER", "J+10 - J+20",
     "Service aides HUARD (referent administratif)",
     "Mandataire CEE / EDF OA / ADEME / Bpifrance / Anah / Avere\nClient (signature mandat)",
     "1. Signature mandat client (cession CEE, mandat ADVENIR, mandat Anah copro, etc.)\n2. Constitution dossier : devis date et signe ulterieurement, attestations RGE, audit/note de calcul, photos avant travaux, releves de compteurs, plans, references cadastrales\n3. Pour copro : vote AG, mandat syndic, PV AG\n4. Pour PV : declaration prealable, attestation Consuel, contrat raccordement Enedis",
     "Dossier complet selon cahier des charges du dispositif\nMandats signes\nAttestations sur l'honneur\nPhotos georeferences",
     "Toute piece manquante = rejet du dossier ou versement reduit. Checklist par dispositif obligatoire."),

    ("PHASE 4\nENGAGEMENT / DEPOT", "J+15 - J+25",
     "Service aides HUARD",
     "Plateformes des financeurs : EMMY (CEE), ADVENIR, MyADEME, EDF OA, Anah/MPR, Bpifrance, OPERAT",
     "1. Depot du dossier sur la plateforme du dispositif\n2. Generation de la date d'engagement CEE\n3. Apres l'attestation d'engagement : signature devis avec le client\n4. Notification au client de la prime estimee et du calendrier",
     "Recepisse d'engagement CEE (ETOILE / SIMUL)\nNotification d'eligibilite ADVENIR\nAccord prealable Anah\nDevis signe par client (date posterieure)",
     "ORDRE STRICT : engagement AVANT signature devis pour CEE et ADVENIR. Inversion = nullite."),

    ("PHASE 5\nREALISATION DES TRAVAUX", "J+25 - J+150",
     "Conduite de travaux HUARD + Atelier",
     "Sous-traitants RGE\nBureau de controle si requis (Consuel, AC)",
     "1. Execution conforme au dossier d'engagement\n2. Tracabilite : fiches techniques produits, certificats, lots\n3. Photos avant/pendant/apres georeferences\n4. PV de pose et OPR contradictoires\n5. Visite de controle ADEME / mandataire (echantillon)",
     "Factures detaillees mentionnant marque/modele/numero de serie\nPV de pose, OPR\nBordereau de suivi des dechets (BSD)\nFiches techniques + DoP des equipements\nCertificats RGE valides a la date des travaux\nPour PV : Consuel + contrat OA",
     "Factures detaillees obligatoires (CEE refuse les libelles generiques). Conservation 10 ans (controle a posteriori)."),

    ("PHASE 6\nVERSEMENT & SUIVI", "J+150 - J+330",
     "Service aides HUARD + Comptabilite",
     "Mandataire CEE - ADEME - Anah - Avere - EDF OA - Bpifrance",
     "1. Depot des pieces de cloture sur les plateformes\n2. Reception attestation de fin de travaux\n3. Versement de la prime (au client ou a HUARD en cession)\n4. Reedition factures avec mention 'finance par CEE n XXXX' si requis\n5. Suivi des controles a posteriori (3-5 ans selon dispositif)",
     "Attestation de fin de travaux signee client\nPV reception\nFiches valorisees CEE\nReleve d'encaissement",
     "Delai moyen versement : CEE 3-6 mois - ADVENIR 4-8 mois - Fonds Chaleur 6-12 mois - MPR Copro 6-9 mois. Provisionner la tresorerie en cas d'avance."),
]

# Couleurs par phase
phase_colors = ["F4A261", "E76F51", "5D81A6", "2A9D8F", "82C341", "1F2F4D"]

for i, phase in enumerate(phases):
    excel_row = 5 + i
    for ci, val in enumerate(phase, start=1):
        c = ws10.cell(row=excel_row, column=ci, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        c.font = Font(size=10)
    # Phase column
    ws10.cell(row=excel_row, column=1).fill = PatternFill("solid", fgColor=phase_colors[i])
    ws10.cell(row=excel_row, column=1).font = Font(size=11, bold=True, color="FFFFFF")
    ws10.cell(row=excel_row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Critical column
    ws10.cell(row=excel_row, column=7).fill = PatternFill("solid", fgColor="FFE9C4")
    ws10.cell(row=excel_row, column=7).font = Font(size=10, bold=True, color="8B4500")
    ws10.row_dimensions[excel_row].height = 145

set_widths(ws10, [18, 14, 22, 26, 50, 38, 38])

# Bloc RACI sous le tableau
raci_row_start = 5 + len(phases) + 2
ws10.cell(row=raci_row_start, column=1, value="MATRICE RACI - QUI FAIT QUOI EN INTERNE HUARD").font = Font(size=14, bold=True, color="FFFFFF")
ws10.cell(row=raci_row_start, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws10.merge_cells(start_row=raci_row_start, start_column=1, end_row=raci_row_start, end_column=7)
ws10.row_dimensions[raci_row_start].height = 26

raci_headers = ["Action", "Commercial", "Charge d'affaires", "Service aides", "Conduite travaux", "Comptabilite", "Direction"]
header_row(ws10, raci_row_start + 1, raci_headers, fill=BLUE)
ws10.row_dimensions[raci_row_start + 1].height = 32

raci = [
    ("Detection opportunite / qualification client", "R", "C", "I", "I", "-", "I"),
    ("Pre-etude technique + chiffrage devis", "C", "R", "C", "C", "-", "-"),
    ("Calcul aides + choix du dispositif", "C", "C", "R", "-", "-", "I"),
    ("Constitution dossier administratif", "C", "C", "R", "I", "-", "-"),
    ("Mandat client / signature", "R", "C", "C", "-", "-", "-"),
    ("Engagement plateformes (CEE, ADVENIR...)", "I", "C", "R", "-", "-", "-"),
    ("Signature devis (apres engagement)", "R", "C", "I", "-", "-", "I si > 100k"),
    ("Execution travaux + tracabilite pieces", "I", "C", "I", "R", "-", "-"),
    ("Depot pieces de cloture", "I", "C", "R", "C", "-", "-"),
    ("Versement aide + facturation finale", "I", "I", "R", "-", "C", "-"),
    ("Suivi controle a posteriori (3-5 ans)", "-", "-", "R", "C", "I", "-"),
]

for i, line in enumerate(raci):
    excel_row = raci_row_start + 2 + i
    for ci, val in enumerate(line, start=1):
        c = ws10.cell(row=excel_row, column=ci, value=val)
        c.border = border
        c.font = Font(size=10, bold=(ci == 1))
        c.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center", wrap_text=True)
        if ci == 1:
            c.font = Font(size=10, bold=True, color=NAVY)
        else:
            color_map = {"R": (GREEN, "FFFFFF"), "A": (ORANGE := "F4A261", "FFFFFF"), "C": ("FFE9C4", NAVY), "I": (LIGHT, NAVY), "-": ("FFFFFF", GREY)}
            v = val if val in color_map else val.split()[0] if val.split()[0] in color_map else "-"
            fill_c, font_c = color_map.get(v, (LIGHT, NAVY))
            c.fill = PatternFill("solid", fgColor=fill_c)
            c.font = Font(size=10, bold=True, color=font_c if isinstance(font_c, str) else "FFFFFF")
    ws10.row_dimensions[excel_row].height = 22

# Legende
leg_row = raci_row_start + 2 + len(raci) + 1
ws10.cell(row=leg_row, column=1, value="Legende RACI :").font = Font(bold=True, color=NAVY, size=10)
ws10.cell(row=leg_row, column=2, value="R = Realise").fill = PatternFill("solid", fgColor=GREEN)
ws10.cell(row=leg_row, column=2).font = Font(color="FFFFFF", bold=True, size=10)
ws10.cell(row=leg_row, column=2).alignment = Alignment(horizontal="center")
ws10.cell(row=leg_row, column=3, value="C = Consulte").fill = PatternFill("solid", fgColor="FFE9C4")
ws10.cell(row=leg_row, column=3).font = Font(color=NAVY, bold=True, size=10)
ws10.cell(row=leg_row, column=3).alignment = Alignment(horizontal="center")
ws10.cell(row=leg_row, column=4, value="I = Informe").fill = PatternFill("solid", fgColor=LIGHT)
ws10.cell(row=leg_row, column=4).font = Font(color=NAVY, bold=True, size=10)
ws10.cell(row=leg_row, column=4).alignment = Alignment(horizontal="center")
ws10.cell(row=leg_row, column=5, value="- = Non implique").font = Font(color=GREY, size=10)
ws10.cell(row=leg_row, column=5).alignment = Alignment(horizontal="center")

# Bloc outils & SI a mettre en place
si_row = leg_row + 2
ws10.cell(row=si_row, column=1, value="OUTILS & SYSTEMES D'INFORMATION A METTRE EN PLACE").font = Font(size=14, bold=True, color="FFFFFF")
ws10.cell(row=si_row, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws10.merge_cells(start_row=si_row, start_column=1, end_row=si_row, end_column=7)
ws10.row_dimensions[si_row].height = 26

outils = [
    ("Mandataire CEE partenaire (contrat-cadre)", "Negocier le rachat des kWh cumac (cours +/- spot marche, 7-10 EUR/MWh cumac en 2026). Choisir un mandataire avec interface web pour devis simulateur en temps reel."),
    ("Labellisation ADVENIR (Avere-France)", "Formation 1 jour + audit qualite. Indispensable pour vendre IRVE avec aide. Cout faible, ROI immediat sur le 1er chantier."),
    ("Simulateur CEE / ADVENIR / Fonds Chaleur interne", "Tableur ou outil web mandataire pour chiffrage commercial en 5 min. Le SIMULATEUR ci-joint (onglet 12) couvre les principaux dispositifs."),
    ("CRM + module 'Aide & financement'", "Champ obligatoire 'aides identifiees' sur chaque devis. KPI mensuel : taux de devis comportant des aides chiffrees."),
    ("GED dossier client (drive partage)", "Arborescence type : 01 Qualification - 02 Devis - 03 Mandats - 04 Engagement - 05 Travaux - 06 Cloture - 07 Versement. Conservation 10 ans."),
    ("Tableau de bord commercial 'aides'", "Suivi mensuel : nombre dossiers en cours, montant aides engagees, montant verse, delais moyens, taux de rejet."),
    ("Partenariat BE thermique (audit DEET / Decret tertiaire)", "Convention de sous-traitance ou mandat. Audit OPQIBI 1905 obligatoire pour Coup de pouce renovation tertiaire."),
    ("Convention banque / leasing pour avance de tresorerie", "Pour les clients qui ne peuvent pas attendre 6-12 mois le versement : avance HUARD ou leasing dedie."),
]

header_row(ws10, si_row + 1, ["Outil / dispositif", "Description / mise en place", "", "", "", "", ""], fill=BLUE)
ws10.merge_cells(start_row=si_row + 1, start_column=2, end_row=si_row + 1, end_column=7)
ws10.row_dimensions[si_row + 1].height = 26

for i, (nom, desc) in enumerate(outils):
    er = si_row + 2 + i
    c1 = ws10.cell(row=er, column=1, value=nom)
    c1.font = Font(bold=True, color=NAVY, size=10)
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c1.border = border
    c2 = ws10.cell(row=er, column=2, value=desc)
    c2.font = Font(size=10)
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    c2.border = border
    ws10.merge_cells(start_row=er, start_column=2, end_row=er, end_column=7)
    if i % 2 == 0:
        for col in range(1, 8):
            ws10.cell(row=er, column=col).fill = PatternFill("solid", fgColor=LIGHT)
    ws10.row_dimensions[er].height = 38


# =========================================================================
# 11. MATRICE PROJETS X AIDES
# =========================================================================
ws11 = wb.create_sheet("10 - Matrice projet x aides")
ws11.sheet_view.showGridLines = False
title_cell(ws11, "A1", "MATRICE PROJET x AIDES MOBILISABLES", size=16)
ws11.merge_cells("A1:J1")
ws11.row_dimensions[1].height = 30
ws11["A2"] = "Pour chaque type de projet HUARD : tous les dispositifs cumulables - taux de couverture moyen - bases de calcul"
ws11["A2"].font = Font(italic=True, color=GREY, size=10)
ws11.merge_cells("A2:J2")

mat_headers = ["#", "Type de projet HUARD", "Activite", "Cible client", "CEE / Coup de pouce", "Subvention (ADEME / Anah / Region)",
               "Tarif / Prime PV / IRVE", "Pret bonifie", "Taux de couverture moyen", "Reste a charge typique"]
header_row(ws11, 4, mat_headers, fill=NAVY)
ws11.row_dimensions[4].height = 40

projets = [
    (1, "Relamping LED tertiaire + gestion", "ELEC", "Bureaux, commerces, parking couvert",
     "CEE BAT-EQ-127 + BAT-EQ-133 (gestion)", "Tremplin PME ADEME (si TPE/PME)", "-", "PEE Bpifrance",
     "30 a 70%", "30-70% du HT"),
    (2, "Remplacement chaudiere fioul / gaz par PAC (collectif copro)", "CVC", "Copropriete chauffage collectif",
     "CEE BAT-TH-104 + Coup de pouce 'Chauffage tertiaire' bonifie x2-x4", "MaPrimeRenov Copro (jusqu'a 25%) + Energies POSIT'IF (tiers-financement)", "-", "Eco-PTZ copro (50k/log) + Pret Vert",
     "40 a 80%", "20-60% du HT (souvent 0 cash via tiers-financement)"),
    (3, "PAC tertiaire (>100 kW thermique)", "CVC", "Tertiaire DEET",
     "CEE BAT-TH-104 + Coup de pouce", "Fonds Chaleur si EnR (geothermie)", "-", "Pret Vert Bpifrance",
     "30 a 65%", "35-70% du HT"),
    (4, "Installation PV autoconsommation 9-100 kWc", "ELEC", "PME tertiaire / industrie / agricole",
     "-", "Region IDF (selon AAP)", "Prime autoconso S26 (5 ans) + Tarif EDF OA 20 ans surplus", "Pret Vert / PEE",
     "15 a 25% (prime) + 60% revenus rachat sur 20 ans", "75-85% en capex - ROI 6-9 ans"),
    (5, "Ombrieres PV parking (Loi APER) 100-500 kWc", "ELEC", "Hyper / logistique / industrie",
     "-", "Region IDF + Fonds Chaleur si associee a EnR", "Prime autoconso (jusqu'a 100 kWc) + Tarif rachat 20 ans (ou AO CRE > 500 kWc)", "Pret Vert (gros volume)",
     "10 a 30% en capex direct + revenus 20 ans", "Investissement amorti 7-10 ans"),
    (6, "GTB classe A/B (decret BACS)", "ELEC/CVC", "Tertiaire CVC > 290 kW (2025), > 70 kW (2027)",
     "CEE BAT-TH-116 (gros volume) + Coup de pouce 'Pilotage connecte chauffage'", "Programme ACTEE (collectivites)", "-", "PEE",
     "50 a 80%", "20-50% du HT - obligatoire reglementairement"),
    (7, "Bornes IRVE parking entreprise (10 a 50 points 22 kW)", "ELEC", "Tertiaire / industrie avec parking salaries",
     "-", "Region IDF (selon AAP)", "ADVENIR jusqu'a 1700 EUR HT/point salaries (parking entreprise)", "PEE",
     "30 a 50%", "50-70% du HT + amortissement fiscal"),
    (8, "Bornes IRVE collective copropriete", "ELEC", "Copropriete",
     "-", "MaPrimeRenov Copro (pour parties communes electriques)", "ADVENIR jusqu'a 1660 EUR/point + 3000 EUR pre-equipement collectif", "Eco-PTZ copro",
     "40 a 60%", "Souvent < 500 EUR par coproprietaire equipe"),
    (9, "Raccordement reseau de chaleur EnR&R", "CVC", "Copro / tertiaire en zone urbaine",
     "CEE BAT-TH-127 + Coup de pouce raccordement", "Fonds Chaleur ADEME (selon reseau) + MaPrimeRenov Copro", "-", "Eco-PTZ copro",
     "40 a 70%", "30-60% du HT"),
    (10, "Audit + bouquet renovation tertiaire (>= 30% gain)", "MULTI", "Tertiaire DEET > 1000 m2",
     "Coup de pouce 'Renovation performante tertiaire' (bonification jusqu'a x4)", "Region IDF + ADEME (selon AAP) + Diag Decarbon'Action Bpifrance", "-", "Pret Vert (gros)",
     "40 a 70%", "30-60% du HT - mise en conformite DEET"),
    (11, "Calorifugeage chaufferie collective", "CVC/MAINT", "Copro / tertiaire avec chaufferie",
     "CEE BAT-TH-145 + BAT-TH-146", "-", "-", "-",
     "80 a 100%", "Souvent 0 EUR pour le client"),
    (12, "VMC double flux ecole / bureau", "CVC", "Tertiaire / etablissement scolaire",
     "CEE BAT-TH-125 / 155", "Programme ACTEE (ecoles publiques) + Region IDF + AIRPARIF (QAI)", "-", "PEE",
     "25 a 50%", "50-75% du HT"),
    (13, "Variateurs vitesse sur moteurs industriels", "ELEC/MAINT", "Industrie - sites avec moteurs > 0.75 kW charge variable",
     "CEE IND-UT-103", "Tremplin PME ADEME + Diag Eco-Flux Bpifrance", "-", "PEE",
     "30 a 60%", "40-70% du HT - ROI < 3 ans"),
    (14, "Recuperation chaleur fatale (compresseur, groupe froid)", "CVC/MAINT", "Industrie / datacenter",
     "CEE IND-UT-117 / IND-UT-102", "Fonds Chaleur ADEME + Decarbonation industrie (France 2030)", "-", "Pret Vert",
     "40 a 75%", "25-60% du HT"),
    (15, "Remplacement chaudiere individuelle par PAC (maison)", "CVC", "Particulier proprietaire occupant",
     "CEE BAR-TH-104 + Coup de pouce chauffage", "MaPrimeRenov par geste ou accompagne", "-", "Eco-PTZ",
     "40 a 90% (selon revenus)", "10-60% du HT + TVA 5,5%"),
    (16, "Datacenter - efficacite refroidissement", "INFO/CVC", "Hebergeurs, salles serveurs entreprise",
     "CEE IND-UT-102 + fiches specifiques", "Fonds Chaleur (chaleur fatale) + France 2030", "-", "Pret Vert",
     "25 a 50%", "50-75% du HT"),
]

cat_color_map = {"ELEC": "DCE7F3", "CVC": "D9F0E3", "ELEC/CVC": "C7D9EC", "CVC/MAINT": "D9F0E3", "ELEC/MAINT": "DCE7F3", "INFO/CVC": "EAD9F0", "MULTI": "FFE9C4"}

for i, projet in enumerate(projets):
    er = 5 + i
    for ci, val in enumerate(projet, start=1):
        c = ws11.cell(row=er, column=ci, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        c.font = Font(size=10)
        if ci == 1:
            c.font = Font(size=11, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 2:
            c.font = Font(size=10, bold=True, color=NAVY)
        elif ci == 3:
            cat = str(val)
            c.fill = PatternFill("solid", fgColor=cat_color_map.get(cat, LIGHT))
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=10, bold=True, color=NAVY)
        elif ci == 9:
            c.font = Font(size=10, bold=True, color=GREEN)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 10:
            c.font = Font(size=10, bold=True, color="B85C00")
        if i % 2 == 0 and ci not in (1, 3):
            c.fill = PatternFill("solid", fgColor=LIGHT)
    ws11.row_dimensions[er].height = 70

set_widths(ws11, [5, 32, 11, 24, 32, 32, 36, 16, 16, 26])
ws11.auto_filter.ref = f"A4:J{4 + len(projets)}"


# =========================================================================
# 12. BASES DE CALCUL - FORMULES PAR DISPOSITIF
# =========================================================================
ws12 = wb.create_sheet("11 - Bases de calcul")
ws12.sheet_view.showGridLines = False
title_cell(ws12, "A1", "BASES DE CALCUL DES AIDES - FORMULES & EXEMPLES", size=16)
ws12.merge_cells("A1:F1")
ws12.row_dimensions[1].height = 30
ws12["A2"] = "Pour chaque dispositif : formule, parametres a recuperer, exemple chiffre - aide le commercial a estimer en 5 min."
ws12["A2"].font = Font(italic=True, color=GREY, size=10)
ws12.merge_cells("A2:F2")

bc_headers = ["Dispositif", "Formule de calcul", "Parametres d'entree", "Unite de sortie", "Exemple chiffre", "Source du parametre"]
header_row(ws12, 4, bc_headers, fill=NAVY)
ws12.row_dimensions[4].height = 36

bc_data = [
    ("CEE standard (toutes fiches BAT/IND/BAR)",
     "Prime = (kWh cumac de la fiche) x (cours marche kWh cumac) / 1000",
     "kWh cumac : depend de la fiche + zone climatique + duree usage\nCours marche : ~7-10 EUR / MWh cumac (T2 2026, indicatif)",
     "EUR HT",
     "BAT-EQ-127 luminaire 60W remplace par 30W LED, zone H1, 14h/j : 8 MWh cumac x 8,5 EUR/MWh = 68 EUR / luminaire",
     "https://emmy.atee.fr (cours) + arrete CEE (kWh cumac)"),

    ("CEE bonifie - Coup de pouce",
     "Prime = (kWh cumac fiche) x (multiplicateur 2 a 4) x cours / 1000",
     "Multiplicateur : depend du Coup de pouce (sortie fioul x4, sortie gaz x2, etc.)",
     "EUR HT",
     "Coup de pouce chauffage tertiaire sortie fioul, multiplicateur x4 : prime = prime CEE de base x 4",
     "Arrete coup de pouce en vigueur"),

    ("Prime a l'autoconsommation PV (arrete S26)",
     "Prime = (EUR / kWc selon tranche) x (P installee en kWc)",
     "Tranche puissance + bareme trimestriel CRE",
     "EUR (verse en 5 annuites)",
     "PV 50 kWc : 50 x 100 EUR/kWc = 5 000 EUR sur 5 ans (T2 2026 indicatif)",
     "Arrete tarifaire S26 - revision trimestrielle CRE"),

    ("Tarif d'achat PV (EDF OA) - surplus",
     "Revenu annuel = (kWh vendus) x (tarif EUR/kWh)",
     "Production annuelle estimee (kWh) x part de surplus (typique 60-80%) x tarif trimestriel",
     "EUR / an (contrat 20 ans)",
     "PV 100 kWc Ile-de-France : ~110 000 kWh/an x 70% surplus x 0,1297 EUR/kWh = 9 980 EUR/an",
     "EDF OA - arrete tarifaire S26"),

    ("Prime ADVENIR - bornes IRVE",
     "Aide = MIN(% HT du cout d'installation ; plafond EUR par point)",
     "Type de parking + puissance borne + nombre de points",
     "EUR HT",
     "10 bornes 22 kW parking entreprise salaries : 10 x MIN(50% HT ; 1 700 EUR) = jusqu'a 17 000 EUR HT",
     "Avere-France ADVENIR + cahier des charges en cours"),

    ("Fonds Chaleur ADEME - mode forfaitaire",
     "Aide = (MWh EnR produits / an) x (forfait EUR / MWh sur 20 ans)",
     "Production EnR annuelle estimee + forfait selon energie (geothermie 80 EUR/MWh, biomasse 25-35 EUR/MWh)",
     "EUR (capitalise sur 20 ans, verse a l'investissement)",
     "Chaufferie biomasse 500 kW, 1500 MWh EnR / an : 1500 x 30 EUR = 45 000 EUR / an x 20 ans = 900 000 EUR (capitalise selon regle ADEME)",
     "Guide Fonds Chaleur ADEME + arretes annuels"),

    ("Fonds Chaleur ADEME - mode aide a l'investissement",
     "Aide = % du capex eligible (variable selon taille entreprise et type)",
     "Cout eligible HT + taux selon AFR + taille",
     "EUR HT",
     "Geothermie sondes 200 kW cout eligible 250 000 EUR HT - PME zone non AFR : 45% = 112 500 EUR",
     "Encadrement aides d'Etat + guide ADEME"),

    ("MaPrimeRenov Copropriete",
     "Aide = (% selon gain energetique) x (cout HT plafonne par logement)",
     "Nombre logements + gain energetique theorique + plafond travaux par logement (25 000 EUR HT/log)",
     "EUR HT",
     "Copro 50 logements, gain 50% : 25% x 25 000 EUR x 50 = 312 500 EUR + bonus sortie passoire",
     "Decrets Anah - bareme MPR Copro 2026"),

    ("MaPrimeRenov individuelle (parcours accompagne)",
     "Aide = % HT plafonne selon revenus + gain energetique",
     "Revenu fiscal de reference + gain energetique vise (35 / 50 / 65% saut classe)",
     "EUR HT",
     "Menage modeste, renovation gain 65% : 80% HT plafonne 70 000 EUR = jusqu'a 56 000 EUR",
     "France Renov + bareme Anah 2026"),

    ("Eco-PTZ",
     "Pret a 0% - montant plafonne selon nombre de travaux",
     "Type et nombre de travaux RGE",
     "EUR (pret amortissable jusqu'a 20 ans)",
     "Bouquet 3 travaux ou + : pret jusqu'a 50 000 EUR a 0% sur 20 ans",
     "Banque partenaire Eco-PTZ"),

    ("TVA reduite 5,5%",
     "Economie = (HT) x (20% - 5,5%) = (HT) x 14,5%",
     "Montant HT travaux eligibles + attestation client",
     "EUR",
     "Devis 10 000 EUR HT travaux PAC residentiel : economie TVA = 1 450 EUR",
     "Art. 278-0 bis A CGI"),

    ("TVA reduite 10%",
     "Economie = (HT) x (20% - 10%) = (HT) x 10%",
     "Montant HT travaux eligibles + attestation client",
     "EUR",
     "Devis 10 000 EUR HT travaux d'amelioration : economie TVA = 1 000 EUR",
     "Art. 279-0 bis CGI"),

    ("Bpifrance - Pret Economies d'Energie (PEE)",
     "Pret bonifie - montant 10 a 500 k EUR - taux fixe ~3-4% (T2 2026)",
     "PME > 3 ans + projet eligible CEE",
     "EUR (pret 3 a 7 ans)",
     "PME industrielle, projet eclairage + variateurs 150 000 EUR : PEE 150 000 EUR sur 5 ans a ~3,5% taux fixe",
     "Bpifrance"),

    ("Tremplin pour la transition ecologique PME (ADEME)",
     "Forfait par action eligible (catalogue ~50 actions)",
     "Type d'action + montant capex",
     "EUR (1 000 a 200 000 EUR total cumule)",
     "Eclairage LED PME : forfait 1 500 EUR + isolation toiture 2 000 EUR + diagnostic 1 000 EUR = 4 500 EUR",
     "ADEME - agir transition"),

    ("Region IDF - aide entreprises TEE",
     "Subvention selon AAP (% du HT ou forfait)",
     "Projet en IDF + PME/ETI + montant eligible HT",
     "EUR HT",
     "Variable selon AAP - typiquement 20-40% HT plafonne (de minimis 200 k EUR sur 3 ans)",
     "iledefrance.fr - AAP en cours"),
]

for i, line in enumerate(bc_data):
    er = 5 + i
    for ci, val in enumerate(line, start=1):
        c = ws12.cell(row=er, column=ci, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        c.font = Font(size=10)
        if ci == 1:
            c.font = Font(size=10, bold=True, color=NAVY)
        elif ci == 5:
            c.fill = PatternFill("solid", fgColor="FFF4B8")
            c.font = Font(size=10, italic=True, color=NAVY)
        if i % 2 == 0 and ci not in (1, 5):
            c.fill = PatternFill("solid", fgColor=LIGHT)
    ws12.row_dimensions[er].height = 75

set_widths(ws12, [32, 38, 34, 18, 42, 26])


# =========================================================================
# 13. SIMULATEURS INTERACTIFS (avec formules Excel vivantes)
# =========================================================================
ws13 = wb.create_sheet("12 - Simulateurs")
ws13.sheet_view.showGridLines = False
title_cell(ws13, "A1", "SIMULATEURS INTERACTIFS - CHIFFRAGE EXPRESS DES AIDES", size=16)
ws13.merge_cells("A1:F1")
ws13.row_dimensions[1].height = 30
ws13["A2"] = "Modifier UNIQUEMENT les cases JAUNES - les cases vertes se recalculent automatiquement."
ws13["A2"].font = Font(italic=True, color=GREY, size=10)
ws13.merge_cells("A2:F2")

YELLOW = "FFF4B8"
GREEN_BG = "D9F0E3"

def sim_input(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=YELLOW)
    c.font = Font(bold=True, size=11, color=NAVY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border
    return c

def sim_output(ws, row, col, formula, fmt="#,##0 EUR"):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = PatternFill("solid", fgColor=GREEN_BG)
    c.font = Font(bold=True, size=11, color=GREEN)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border
    c.number_format = fmt
    return c

def sim_label(ws, row, col, text, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, size=11, color=NAVY)
    c.alignment = Alignment(vertical="center")
    c.border = border
    return c

# --- SIMULATEUR 1 : CEE relamping LED bureaux ---
r = 4
ws13.cell(row=r, column=1, value="SIMULATEUR 1 - RELAMPING LED BUREAUX (CEE BAT-EQ-127 + BAT-EQ-133)").font = Font(size=13, bold=True, color="FFFFFF")
ws13.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws13.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws13.row_dimensions[r].height = 26

r = 5
sim_label(ws13, r, 1, "Nombre de luminaires LED remplaces")
sim_input(ws13, r, 2, 500)
sim_label(ws13, r, 3, "kWh cumac par luminaire (zone H1, 14h/j)", bold=False)
sim_input(ws13, r, 4, 8000)
sim_label(ws13, r, 5, "Cours kWh cumac (EUR/MWh cumac)", bold=False)
sim_input(ws13, r, 6, 8.5)

r = 6
sim_label(ws13, r, 1, "Cout total HT du chantier (EUR)")
sim_input(ws13, r, 2, 95000)
sim_label(ws13, r, 3, "Bonification (1 = standard, 2 = coup de pouce)")
sim_input(ws13, r, 4, 1)

r = 7
sim_label(ws13, r, 1, "Prime CEE estimee (EUR HT)")
sim_output(ws13, r, 2, "=B5*D5*F5*D6/1000")
sim_label(ws13, r, 3, "% de couverture du chantier")
sim_output(ws13, r, 4, "=B7/B6", fmt="0.0%")

r = 8
sim_label(ws13, r, 1, "Reste a charge client (EUR HT)")
sim_output(ws13, r, 2, "=B6-B7")
sim_label(ws13, r, 3, "ROI (annees) - hypothese 30% econo / an")
sim_output(ws13, r, 4, "=B8/(B6*0.3)", fmt="0.0 \" ans\"")

# --- SIMULATEUR 2 : PV autoconsommation ---
r = 11
ws13.cell(row=r, column=1, value="SIMULATEUR 2 - PV AUTOCONSOMMATION (prime S26 + tarif EDF OA)").font = Font(size=13, bold=True, color="FFFFFF")
ws13.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws13.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws13.row_dimensions[r].height = 26

r = 12
sim_label(ws13, r, 1, "Puissance installee (kWc)")
sim_input(ws13, r, 2, 100)
sim_label(ws13, r, 3, "Prime EUR / kWc (selon tranche, T2 2026 indicatif)", bold=False)
sim_input(ws13, r, 4, 100)
sim_label(ws13, r, 5, "Production annuelle (kWh / kWc)", bold=False)
sim_input(ws13, r, 6, 1100)

r = 13
sim_label(ws13, r, 1, "Cout total installation HT (EUR)")
sim_input(ws13, r, 2, 110000)
sim_label(ws13, r, 3, "Part autoconsommee (typique 30-60%)")
sim_input(ws13, r, 4, 0.4)
sim_label(ws13, r, 5, "Tarif rachat surplus (EUR/kWh)", bold=False)
sim_input(ws13, r, 6, 0.1297)

r = 14
sim_label(ws13, r, 1, "Prix electricite achetee (EUR/kWh)")
sim_input(ws13, r, 2, 0.21)

r = 15
sim_label(ws13, r, 1, "Prime totale autoconso (verse 5 ans)")
sim_output(ws13, r, 2, "=B12*D12")
sim_label(ws13, r, 3, "Production annuelle (kWh)")
sim_output(ws13, r, 4, "=B12*F12", fmt="#,##0 \" kWh\"")

r = 16
sim_label(ws13, r, 1, "Revenu rachat surplus / an")
sim_output(ws13, r, 2, "=D15*(1-D13)*F13")
sim_label(ws13, r, 3, "Economie autoconso / an")
sim_output(ws13, r, 4, "=D15*D13*B14")

r = 17
sim_label(ws13, r, 1, "Gain annuel total (revenu + econo)")
sim_output(ws13, r, 2, "=B16+D16")
sim_label(ws13, r, 3, "Retour sur investissement (annees)")
sim_output(ws13, r, 4, "=(B13-B15)/B17", fmt="0.0 \" ans\"")

r = 18
sim_label(ws13, r, 1, "Gain cumule 20 ans (- prime deduite)")
sim_output(ws13, r, 2, "=B17*20+B15")
sim_label(ws13, r, 3, "VAN brute 20 ans")
sim_output(ws13, r, 4, "=B18-B13")

# --- SIMULATEUR 3 : Bornes IRVE ADVENIR ---
r = 21
ws13.cell(row=r, column=1, value="SIMULATEUR 3 - BORNES IRVE (ADVENIR - parking entreprise)").font = Font(size=13, bold=True, color="FFFFFF")
ws13.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws13.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws13.row_dimensions[r].height = 26

r = 22
sim_label(ws13, r, 1, "Nombre de points de charge")
sim_input(ws13, r, 2, 10)
sim_label(ws13, r, 3, "Cout HT par point (installation incluse)", bold=False)
sim_input(ws13, r, 4, 3500)
sim_label(ws13, r, 5, "Plafond ADVENIR EUR / point (T2 2026)", bold=False)
sim_input(ws13, r, 6, 1700)

r = 23
sim_label(ws13, r, 1, "Taux ADVENIR (% du HT)")
sim_input(ws13, r, 2, 0.5)

r = 24
sim_label(ws13, r, 1, "Cout total HT")
sim_output(ws13, r, 2, "=B22*D22")
sim_label(ws13, r, 3, "Aide ADVENIR par point")
sim_output(ws13, r, 4, "=MIN(D22*B23,F22)")

r = 25
sim_label(ws13, r, 1, "Aide ADVENIR totale")
sim_output(ws13, r, 2, "=B22*D24")
sim_label(ws13, r, 3, "Reste a charge client HT")
sim_output(ws13, r, 4, "=B24-B25")

r = 26
sim_label(ws13, r, 1, "% de couverture")
sim_output(ws13, r, 2, "=B25/B24", fmt="0.0%")

# --- SIMULATEUR 4 : Remplacement chaudiere copro par PAC ---
r = 29
ws13.cell(row=r, column=1, value="SIMULATEUR 4 - REMPLACEMENT CHAUDIERE FIOUL PAR PAC COLLECTIVE (copro)").font = Font(size=13, bold=True, color="FFFFFF")
ws13.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws13.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws13.row_dimensions[r].height = 26

r = 30
sim_label(ws13, r, 1, "Nombre de logements")
sim_input(ws13, r, 2, 50)
sim_label(ws13, r, 3, "Cout total HT chantier (EUR)", bold=False)
sim_input(ws13, r, 4, 380000)
sim_label(ws13, r, 5, "Gain energetique (>= 35% requis MPR)", bold=False)
sim_input(ws13, r, 6, 0.5)

r = 31
sim_label(ws13, r, 1, "Prime CEE BAT-TH-104 (estimee)")
sim_input(ws13, r, 2, 35000)
sim_label(ws13, r, 3, "Multiplicateur Coup de pouce sortie fioul")
sim_input(ws13, r, 4, 4)

r = 32
sim_label(ws13, r, 1, "Plafond MPR Copro EUR / logement HT")
sim_input(ws13, r, 2, 25000)
sim_label(ws13, r, 3, "Taux MPR (selon gain >= 50%)")
sim_input(ws13, r, 4, 0.35)

r = 33
sim_label(ws13, r, 1, "Prime CEE bonifiee")
sim_output(ws13, r, 2, "=B31*D31")
sim_label(ws13, r, 3, "Aide MPR Copro estimee")
sim_output(ws13, r, 4, "=MIN(B30*B32,D30)*D32")

r = 34
sim_label(ws13, r, 1, "Total aides cumulees")
sim_output(ws13, r, 2, "=B33+D33")
sim_label(ws13, r, 3, "Reste a charge HT")
sim_output(ws13, r, 4, "=D30-B34")

r = 35
sim_label(ws13, r, 1, "% couverture")
sim_output(ws13, r, 2, "=B34/D30", fmt="0.0%")
sim_label(ws13, r, 3, "Reste par logement (HT)")
sim_output(ws13, r, 4, "=D34/B30")

# --- SIMULATEUR 5 : GTB BACS (BAT-TH-116) ---
r = 38
ws13.cell(row=r, column=1, value="SIMULATEUR 5 - GTB CLASSE A (BAT-TH-116 - decret BACS)").font = Font(size=13, bold=True, color="FFFFFF")
ws13.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws13.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws13.row_dimensions[r].height = 26

r = 39
sim_label(ws13, r, 1, "Surface chauffee (m2)")
sim_input(ws13, r, 2, 5000)
sim_label(ws13, r, 3, "Forfait kWh cumac / m2 (classe A bureaux)", bold=False)
sim_input(ws13, r, 4, 1850)
sim_label(ws13, r, 5, "Cours kWh cumac (EUR/MWh)", bold=False)
sim_input(ws13, r, 6, 8.5)

r = 40
sim_label(ws13, r, 1, "Cout total GTB HT (EUR)")
sim_input(ws13, r, 2, 180000)

r = 41
sim_label(ws13, r, 1, "Prime CEE BAT-TH-116 estimee")
sim_output(ws13, r, 2, "=B39*D39*F39/1000")
sim_label(ws13, r, 3, "% couverture")
sim_output(ws13, r, 4, "=B41/B40", fmt="0.0%")

r = 42
sim_label(ws13, r, 1, "Reste a charge HT")
sim_output(ws13, r, 2, "=B40-B41")
sim_label(ws13, r, 3, "ROI (hypothese 25% econo conso CVC + eclairage)")
sim_output(ws13, r, 4, "=B42/(B40*0.25)", fmt="0.0 \" ans\"")

# --- Legende ---
r = 45
ws13.cell(row=r, column=1, value="LEGENDE").font = Font(bold=True, color=NAVY, size=11)
r = 46
c = ws13.cell(row=r, column=1, value="CASE JAUNE = parametre a entrer")
c.fill = PatternFill("solid", fgColor=YELLOW)
c.border = border
c.font = Font(bold=True, color=NAVY, size=10)
c = ws13.cell(row=r, column=3, value="CASE VERTE = calcul automatique")
c.fill = PatternFill("solid", fgColor=GREEN_BG)
c.border = border
c.font = Font(bold=True, color=GREEN, size=10)

r = 47
ws13.cell(row=r, column=1, value="Avertissement : barèmes indicatifs T2 2026. Verifier les cours CEE (EMMY.atee.fr) et tarifs PV (CRE) avant tout engagement client.").font = Font(italic=True, color=GREY, size=10)
ws13.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

set_widths(ws13, [36, 16, 36, 16, 36, 16])
for r in range(5, 43):
    if ws13.cell(row=r, column=1).value is not None:
        ws13.row_dimensions[r].height = 26

# Mettre a jour le sommaire pour pointer vers les nouveaux onglets
ws_sommaire = wb["Sommaire"]

# Demerger l'avertissement avant d'ajouter
for mr in list(ws_sommaire.merged_cells.ranges):
    if mr.min_row >= 23:
        ws_sommaire.unmerge_cells(str(mr))

# Effacer ancien avertissement
for r in range(23, 32):
    for col in range(2, 9):
        cell = ws_sommaire.cell(row=r, column=col)
        cell.value = None
        cell.fill = PatternFill(fill_type=None)

# Ajout des 4 nouveaux onglets dans le sommaire
sommaire_supp = [
    ("09", "Plan administratif aide client", "Pipeline 6 phases + RACI interne + outils/SI a mettre en place"),
    ("10", "Matrice projet x aides", "16 types de projets HUARD avec aides cumulables, couverture moyenne, reste a charge"),
    ("11", "Bases de calcul", "Formules detaillees par dispositif + parametres d'entree + exemples chiffres"),
    ("12", "Simulateurs interactifs", "5 simulateurs Excel vivants : LED, PV, IRVE, PAC copro, GTB BACS - entrer ses valeurs"),
]
for i, (n, t, d) in enumerate(sommaire_supp, start=21):
    c1 = ws_sommaire.cell(row=i, column=2, value=n)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = border
    c2 = ws_sommaire.cell(row=i, column=3, value=t)
    c2.font = Font(bold=True, color=NAVY)
    c2.border = border
    c3 = ws_sommaire.cell(row=i, column=4, value=d)
    c3.alignment = Alignment(vertical="center", wrap_text=True)
    c3.border = border
    ws_sommaire.row_dimensions[i].height = 28

# Avertissement decalé apres
ws_sommaire["B27"] = "AVERTISSEMENT"
ws_sommaire["B27"].font = Font(bold=True, color="FFFFFF")
ws_sommaire["B27"].fill = PatternFill("solid", fgColor=RED)
ws_sommaire.merge_cells("B27:H27")

avert = (
    "Les montants, plafonds et conditions d'eligibilite sont ceux en vigueur au 16/05/2026. "
    "Les baremes CEE (bonifications, coups de pouce), les arretes tarifaires PV (S21/S26) et les enveloppes ADEME "
    "evoluent chaque trimestre voire chaque mois. AVANT TOUT ENGAGEMENT COMMERCIAL CHIFFRE, valider les montants "
    "actualises aupres de : (1) le mandataire CEE partenaire, (2) ADEME / Bpifrance pour les subventions, "
    "(3) la Region IDF pour les aides regionales. Document non contractuel - usage interne strict."
)
ws_sommaire["B28"] = avert
ws_sommaire["B28"].font = Font(size=10, italic=True, color=GREY)
ws_sommaire["B28"].alignment = Alignment(wrap_text=True, vertical="top")
ws_sommaire.merge_cells("B28:H31")
for r in range(28, 32):
    ws_sommaire.row_dimensions[r].height = 22

# Sauvegarde
wb.save(OUT)
print(f"OK -> {OUT}")
